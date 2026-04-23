"""Trades API — list, detail, close, close-all, notes, export."""
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from db.connection import get_session
from db.models import Trade, TradeCommand, Strategy
import io

router = APIRouter(tags=["trades"])


class CloseRequest(BaseModel):
    contracts: Optional[int] = None  # None = close all


def _trade_to_dict(t: Trade) -> dict:
    # Multi-strategy v2: per-leg fields (symbol, direction, prices,
    # contracts, IB identifiers, brackets, strategy levels) now live on
    # trade_legs. For single-leg strategies (everything today) we surface
    # the first leg's fields at trade level to preserve the legacy
    # envelope-shaped response the UI expects.
    leg0 = t.legs[0] if t.legs else None
    return {
        "id": t.id, "account": t.account, "ticker": t.ticker,
        "strategy_id": t.strategy_id,
        "strategy_name": t.strategy.name if t.strategy else None,
        "strategy_display_name": t.strategy.display_name if t.strategy else None,
        "symbol": leg0.symbol if leg0 else None,
        "direction": leg0.direction if leg0 else None,
        "contracts_entered": leg0.contracts_entered if leg0 else None,
        "contracts_open": leg0.contracts_open if leg0 else None,
        "contracts_closed": leg0.contracts_closed if leg0 else None,
        "entry_price": float(leg0.entry_price) if leg0 and leg0.entry_price else None,
        "exit_price": float(leg0.exit_price) if leg0 and leg0.exit_price else None,
        "current_price": float(leg0.current_price) if leg0 and leg0.current_price else None,
        "ib_fill_price": float(leg0.ib_fill_price) if leg0 and leg0.ib_fill_price else None,
        "pnl_pct": float(t.pnl_pct) if t.pnl_pct else 0,
        "pnl_usd": float(t.pnl_usd) if t.pnl_usd else 0,
        "peak_pnl_pct": float(t.peak_pnl_pct) if t.peak_pnl_pct else 0,
        "dynamic_sl_pct": float(t.dynamic_sl_pct) if t.dynamic_sl_pct else 0,
        "profit_target": float(leg0.profit_target) if leg0 and leg0.profit_target else None,
        "stop_loss_level": float(leg0.stop_loss_level) if leg0 and leg0.stop_loss_level else None,
        "signal_type": t.signal_type,
        "entry_time": t.entry_time.isoformat() if t.entry_time else None,
        "exit_time": t.exit_time.isoformat() if t.exit_time else None,
        "status": t.status, "exit_reason": t.exit_reason, "exit_result": t.exit_result,
        "error_message": t.error_message,
        "entry_enrichment": t.entry_enrichment or {},
        "exit_enrichment": t.exit_enrichment or {},
        "notes": t.notes,
        # Bracket visibility (updated by reconcile PASS 4). Now on the leg.
        # Status values: Submitted / PreSubmitted / PendingSubmit
        # (active), Cancelled / ApiCancelled / Inactive / Filled
        # (terminal), MISSING (permId not found on IB at all),
        # NULL (never placed / old row).
        "ib_tp_perm_id":   leg0.ib_tp_perm_id if leg0 else None,
        "ib_sl_perm_id":   leg0.ib_sl_perm_id if leg0 else None,
        "ib_tp_status":    leg0.ib_tp_status if leg0 else None,
        "ib_sl_status":    leg0.ib_sl_status if leg0 else None,
        "ib_tp_price":     float(leg0.ib_tp_price) if leg0 and leg0.ib_tp_price else None,
        "ib_sl_price":     float(leg0.ib_sl_price) if leg0 and leg0.ib_sl_price else None,
        "ib_tp_order_id":  leg0.ib_tp_order_id if leg0 else None,
        "ib_sl_order_id":  leg0.ib_sl_order_id if leg0 else None,
        "ib_brackets_checked_at": (
            leg0.ib_brackets_checked_at.isoformat()
            if leg0 and leg0.ib_brackets_checked_at
            else None
        ),
        # Parent (entry) order IDs for troubleshooting — unique across
        # all IB clients when permId is set. Now on the leg.
        "ib_order_id": leg0.ib_order_id if leg0 else None,
        "ib_perm_id":  leg0.ib_perm_id if leg0 else None,
        "ib_con_id":   leg0.ib_con_id if leg0 else None,
        # Human-readable IB↔DB correlation (TICKER-YYMMDD-NN).
        # Matches IB Order.orderRef / TWS "Order Ref" column.
        "client_trade_id": getattr(t, "client_trade_id", None),
        # Leg count — UI uses this to show an expand caret on multi-leg
        # rows (iron condors, spreads, hedged positions). ENH-047.
        "n_legs": getattr(t, "n_legs", None) or (len(t.legs) if t.legs else 1),
    }


def _leg_to_dict(l) -> dict:
    """Shape one trade_legs row for the UI. Matches ENH-047 drill-down."""
    def _f(v):
        return float(v) if v is not None else None

    sign = 1 if (l.direction or "LONG") == "LONG" else -1
    entry = float(l.entry_price) if l.entry_price is not None else None
    exit_px = float(l.exit_price) if l.exit_price is not None else None
    cur = float(l.current_price) if l.current_price is not None else None
    close_ref = exit_px if exit_px is not None else cur
    per_leg_pnl = None
    if entry is not None and close_ref is not None and l.contracts_entered:
        per_leg_pnl = (
            (close_ref - entry) * l.contracts_entered * (l.multiplier or 100) * sign
        )

    return {
        "leg_id": l.leg_id,
        "trade_id": l.trade_id,
        "leg_index": l.leg_index,
        "leg_role": l.leg_role,
        "sec_type": l.sec_type,
        "symbol": l.symbol,
        "underlying": l.underlying,
        "strike": _f(l.strike),
        "right": l.right,
        "expiry": l.expiry,
        "multiplier": l.multiplier,
        "exchange": l.exchange,
        "currency": l.currency,
        "direction": l.direction,
        "contracts_entered": l.contracts_entered,
        "contracts_open": l.contracts_open,
        "contracts_closed": l.contracts_closed,
        "entry_price": entry,
        "exit_price": exit_px,
        "current_price": cur,
        "ib_fill_price": _f(l.ib_fill_price),
        "profit_target": _f(l.profit_target),
        "stop_loss_level": _f(l.stop_loss_level),
        "ib_order_id": l.ib_order_id,
        "ib_perm_id": l.ib_perm_id,
        "ib_con_id": l.ib_con_id,
        "ib_tp_order_id": l.ib_tp_order_id,
        "ib_tp_perm_id": l.ib_tp_perm_id,
        "ib_tp_status": l.ib_tp_status,
        "ib_tp_price": _f(l.ib_tp_price),
        "ib_sl_order_id": l.ib_sl_order_id,
        "ib_sl_perm_id": l.ib_sl_perm_id,
        "ib_sl_status": l.ib_sl_status,
        "ib_sl_price": _f(l.ib_sl_price),
        "entry_time": l.entry_time.isoformat() if l.entry_time else None,
        "exit_time": l.exit_time.isoformat() if l.exit_time else None,
        "leg_status": l.leg_status,
        "pnl_usd": per_leg_pnl,
    }


@router.get("/trades")
def list_trades(
    status: Optional[str] = None,
    ticker: Optional[str] = None,
    sort: str = "entry_time",
    order: str = "desc",
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
):
    session = get_session()
    if not session:
        raise HTTPException(503, "Database not available")
    try:
        q = session.query(Trade)
        if status:
            q = q.filter(Trade.status == status)
        if ticker:
            q = q.filter(Trade.ticker == ticker.upper())

        # Sorting
        col = getattr(Trade, sort, Trade.entry_time)
        q = q.order_by(col.desc() if order == "desc" else col.asc())

        total = q.count()
        trades = q.offset((page - 1) * limit).limit(limit).all()
        result = [_trade_to_dict(t) for t in trades]
        session.close()
        return {"trades": result, "total": total, "page": page, "limit": limit}
    finally:
        session.close()


@router.get("/trades/{trade_id}")
def get_trade(trade_id: int):
    session = get_session()
    if not session:
        raise HTTPException(503, "Database not available")
    try:
        trade = session.query(Trade).filter(Trade.id == trade_id).first()
        if not trade:
            raise HTTPException(404, "Trade not found")
        result = _trade_to_dict(trade)
        session.close()
        return result
    finally:
        session.close()


@router.get("/trades/{trade_id}/audit")
def get_trade_audit(trade_id: int):
    """Return every system_log row that touched this trade, oldest first.

    Queries ``system_log.details->>'trade_id'`` which is populated by
    ``strategy.audit.log_trade_action``. Also includes any log line
    whose details contain ``from_trade_id`` or ``to_trade_id`` equal
    to the given id, so roll chains are visible from both sides.

    Use this to answer: "who opened this trade? who closed it? what
    reconciled it? no mystery actions."
    """
    session = get_session()
    if session is None:
        raise HTTPException(503, "Database not available")
    try:
        from sqlalchemy import text
        rows = session.execute(text(
            """
            SELECT id, component, level, message, details, created_at
            FROM system_log
            WHERE (details->>'trade_id')::int     = :tid
               OR (details->>'from_trade_id')::int = :tid
               OR (details->>'to_trade_id')::int   = :tid
            ORDER BY created_at ASC
            LIMIT 500
            """
        ), {"tid": trade_id}).fetchall()
        return {
            "trade_id": trade_id,
            "entries": [
                {
                    "id": r[0],
                    "component": r[1],
                    "level": r[2],
                    "message": r[3],
                    "details": r[4] or {},
                    "created_at": r[5].isoformat() if r[5] else None,
                }
                for r in rows
            ],
            "count": len(rows),
        }
    finally:
        session.close()


@router.get("/trades/{trade_id}/legs")
def get_trade_legs(trade_id: int):
    """Return every leg of a trade, ordered by leg_index.

    ENH-047 — Trades-tab per-leg drill-down. Multi-leg trades (iron
    condors, spreads) surface as ONE row in the Trades table; this
    endpoint feeds the expand-panel that shows each leg's symbol,
    direction, fill prices and per-leg P&L.
    """
    session = get_session()
    if session is None:
        raise HTTPException(503, "Database not available")
    try:
        trade = session.query(Trade).filter(Trade.id == trade_id).first()
        if not trade:
            raise HTTPException(404, "Trade not found")
        legs = sorted(list(trade.legs), key=lambda l: l.leg_index)
        return {
            "trade_id": trade_id,
            "ticker": trade.ticker,
            "n_legs": getattr(trade, "n_legs", None) or len(legs),
            "legs": [_leg_to_dict(l) for l in legs],
        }
    finally:
        session.close()


class NotesUpdate(BaseModel):
    notes: str = ""


@router.put("/trades/{trade_id}/notes")
def update_trade_notes(trade_id: int, body: NotesUpdate):
    """Update notes for a trade."""
    session = get_session()
    if not session:
        raise HTTPException(503, "Database not available")
    try:
        trade = session.query(Trade).filter(Trade.id == trade_id).first()
        if not trade:
            raise HTTPException(404, f"Trade {trade_id} not found")
        trade.notes = body.notes
        session.commit()
        return {"status": "ok", "trade_id": trade_id, "notes": body.notes}
    finally:
        session.close()


def _bot_is_running() -> bool:
    """Return True if the bot process is running (DB bot_state.status).
    We queue close commands through the bot when it's up so they go via
    the proper sell-first + pool-aware _atomic_close flow. Only fall
    back to the sidecar's direct-IB close when the bot is actually down
    (emergency path only)."""
    try:
        from db.models import BotState
        session = get_session()
        if not session:
            return False
        state = session.query(BotState).filter(BotState.id == 1).first()
        running = bool(state and state.status == "running")
        session.close()
        return running
    except Exception:
        return False


@router.post("/trades/{trade_id}/close")
async def close_trade(trade_id: int, req: CloseRequest = CloseRequest()):
    """
    Smart close — route depends on bot state:
    1. Bot running → queue a TradeCommand. ExitManager picks it up on
       its next cycle (~1s) and runs it through _atomic_close which
       uses the sell-first + pool-aware bracket cleanup. This is the
       SAFE path — previously we called the sidecar directly and left
       cross-client brackets alive on IB (SPY/GOOGL 2026-04-21).
    2. Bot stopped → fall back to sidecar's emergency close (works,
       but may leave brackets alive because the fresh clientId=99
       connection can't cancel orders owned by the pool's clientIds).
    3. Already-closed row → refuse.
    """
    session = get_session()
    if not session:
        raise HTTPException(503, "Database not available")
    try:
        trade = session.query(Trade).filter(Trade.id == trade_id).first()
        if not trade:
            raise HTTPException(404, "Trade not found")
        if trade.status != "open":
            raise HTTPException(400, f"Trade is already {trade.status}")

        # Preferred path: queue for the bot's exit_manager (proper
        # atomic close with pool-aware bracket cleanup).
        if _bot_is_running():
            cmd = TradeCommand(
                trade_id=trade_id,
                command="close_partial" if req.contracts else "close",
                contracts=req.contracts,
            )
            session.add(cmd)
            session.commit()
            session.close()
            return {"status": "command_queued", "command_id": cmd.id,
                    "note": "Bot exit_manager will process on next cycle "
                            "(atomic close + pool-aware bracket cleanup)"}

        # Bot down — use sidecar's direct-IB close (emergency path).
        # v2: per-leg fields live on trade.legs[0] (single-leg only here).
        leg0 = trade.legs[0] if trade.legs else None
        if leg0 is None:
            raise HTTPException(500, f"Trade {trade_id} has no legs — cannot close")
        try:
            import httpx
            import os
            sidecar_url = os.getenv("BOT_SIDECAR_URL", "http://host.docker.internal:9000")
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.post(f"{sidecar_url}/close-trade",
                    json={"trade_id": trade_id, "symbol": leg0.symbol,
                          "ticker": trade.ticker, "contracts": req.contracts or leg0.contracts_open,
                          "direction": leg0.direction})
                if resp.status_code == 200:
                    result = resp.json()
                    # Update DB based on sidecar response
                    from datetime import datetime, timezone

                    # Use IB fill price, fall back to last known price
                    ib_exit_price = result.get("exit_price", 0)
                    exit_p = float(ib_exit_price) if ib_exit_price else float(leg0.current_price or leg0.entry_price or 0)

                    # Use IB execution time, fall back to now
                    ib_exit_time = result.get("exit_time")
                    if ib_exit_time:
                        try:
                            exit_time = datetime.fromisoformat(str(ib_exit_time).replace('Z', '+00:00'))
                        except Exception:
                            exit_time = datetime.now(timezone.utc)
                    else:
                        exit_time = datetime.now(timezone.utc)

                    if result.get("position_was_open") == False:
                        trade.exit_reason = "CLOSED (BRACKET/IB)"
                    else:
                        trade.exit_reason = "CLOSED (UI)"

                    trade.status = "closed"
                    trade.exit_time = exit_time
                    leg0.exit_price = exit_p
                    leg0.current_price = exit_p  # sync current with exit
                    leg0.exit_time = exit_time
                    leg0.leg_status = "closed"
                    entry = float(leg0.entry_price) if leg0.entry_price else 0
                    trade.pnl_pct = (exit_p - entry) / entry if entry > 0 else 0
                    trade.pnl_usd = (exit_p - entry) * 100 * leg0.contracts_entered
                    trade.exit_result = "WIN" if trade.pnl_pct > 0 else "LOSS" if trade.pnl_pct < 0 else "SCRATCH"
                    leg0.contracts_open = 0
                    leg0.contracts_closed = leg0.contracts_entered
                    session.commit()
                    session.close()
                    return {"status": "closed", "trade_id": trade_id, "detail": result}
        except Exception:
            pass  # Sidecar not available — fall back to command queue

        # Fallback: queue command for bot to process
        cmd = TradeCommand(
            trade_id=trade_id,
            command="close_partial" if req.contracts else "close",
            contracts=req.contracts,
        )
        session.add(cmd)
        session.commit()
        session.close()
        return {"status": "command_queued", "command_id": cmd.id,
                "note": "Bot will process when running"}
    finally:
        session.close()


@router.post("/trades/close-all")
async def close_all_trades():
    """Close all open trades.

    Same routing logic as the single-trade close: queue through the
    bot's exit_manager when the bot is up (sell-first + pool-aware
    bracket cleanup), only fall back to the sidecar when the bot is
    actually down. Previously this always used the sidecar directly,
    which left cross-client brackets alive on IB.
    """
    session = get_session()
    if not session:
        raise HTTPException(503, "Database not available")
    try:
        open_trades = session.query(Trade).filter(Trade.status == "open").all()
        if not open_trades:
            return {"status": "no_open_trades"}

        bot_up = _bot_is_running()
        results = []

        for t in open_trades:
            # Preferred path: queue one TradeCommand per trade; the bot
            # processes them serially through _atomic_close.
            if bot_up:
                cmd = TradeCommand(trade_id=t.id, command="close")
                session.add(cmd)
                results.append({"id": t.id, "status": "queued"})
                continue

            # Bot down — emergency sidecar path (may leave brackets alive).
            # v2: per-leg fields live on t.legs[0] (single-leg only here).
            leg0 = t.legs[0] if t.legs else None
            if leg0 is None:
                results.append({"id": t.id, "status": "skipped_no_legs"})
                continue
            try:
                import httpx
                import os
                sidecar_url = os.getenv("BOT_SIDECAR_URL", "http://host.docker.internal:9000")
                async with httpx.AsyncClient(timeout=15.0) as client:
                    resp = await client.post(f"{sidecar_url}/close-trade",
                        json={"trade_id": t.id, "symbol": leg0.symbol,
                              "ticker": t.ticker, "contracts": leg0.contracts_open,
                              "direction": leg0.direction})
                    if resp.status_code == 200:
                        from datetime import datetime, timezone
                        detail = resp.json()
                        ib_exit = detail.get("exit_price", 0)
                        exit_p = float(ib_exit) if ib_exit else float(leg0.current_price or leg0.entry_price or 0)
                        ib_time = detail.get("exit_time")
                        if ib_time:
                            try:
                                exit_time = datetime.fromisoformat(str(ib_time).replace('Z', '+00:00'))
                            except Exception:
                                exit_time = datetime.now(timezone.utc)
                        else:
                            exit_time = datetime.now(timezone.utc)
                        t.status = "closed"
                        t.exit_time = exit_time
                        leg0.exit_price = exit_p
                        leg0.current_price = exit_p
                        leg0.exit_time = exit_time
                        leg0.leg_status = "closed"
                        t.exit_reason = "CLOSED (UI CLOSE ALL — sidecar)"
                        entry = float(leg0.entry_price) if leg0.entry_price else 0
                        t.pnl_pct = (exit_p - entry) / entry if entry > 0 else 0
                        t.pnl_usd = (exit_p - entry) * 100 * leg0.contracts_entered
                        t.exit_result = "WIN" if t.pnl_pct > 0 else "LOSS" if t.pnl_pct < 0 else "SCRATCH"
                        leg0.contracts_open = 0
                        leg0.contracts_closed = leg0.contracts_entered
                        results.append({"id": t.id, "status": "closed"})
                        continue
            except Exception:
                pass
            # Last-resort: queue even though we couldn't reach sidecar
            cmd = TradeCommand(trade_id=t.id, command="close")
            session.add(cmd)
            results.append({"id": t.id, "status": "queued"})

        session.commit()
        session.close()
        return {"status": "processed", "trades": results, "count": len(results),
                "routed_via": "bot_queue" if bot_up else "sidecar_emergency"}
    finally:
        session.close()


@router.get("/trades/export")
def export_trades(
    status: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
):
    """Export trades to Excel (.xlsx). Supports filtering by status and date range."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    session = get_session()
    if not session:
        raise HTTPException(503, "Database not available")
    try:
        q = session.query(Trade)
        if status:
            q = q.filter(Trade.status == status)
        if start:
            q = q.filter(Trade.entry_time >= start)
        if end:
            q = q.filter(Trade.entry_time <= end + "T23:59:59")
        trades = q.order_by(Trade.entry_time.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"

        # Headers
        headers = ["ID", "Ticker", "Symbol", "Direction", "Contracts", "Entry Price",
                    "Exit Price", "P&L %", "P&L $", "Peak P&L %", "Status",
                    "Exit Reason", "Exit Result", "Signal Type", "Entry Time",
                    "Exit Time", "Notes"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        win_fill = PatternFill("solid", fgColor="C6EFCE")
        loss_fill = PatternFill("solid", fgColor="FFC7CE")
        for row_idx, t in enumerate(trades, 2):
            pnl_pct = float(t.pnl_pct * 100) if t.pnl_pct else 0
            pnl_usd = float(t.pnl_usd) if t.pnl_usd else 0
            peak = float(t.peak_pnl_pct * 100) if t.peak_pnl_pct else 0
            # v2: per-leg fields live on t.legs[0] (single-leg strategies today).
            leg0 = t.legs[0] if t.legs else None
            row_data = [
                t.id, t.ticker,
                leg0.symbol if leg0 else None,
                leg0.direction if leg0 else None,
                leg0.contracts_entered if leg0 else None,
                float(leg0.entry_price) if leg0 and leg0.entry_price else None,
                float(leg0.exit_price) if leg0 and leg0.exit_price else None,
                round(pnl_pct, 2), round(pnl_usd, 2), round(peak, 2),
                t.status, t.exit_reason, t.exit_result, t.signal_type,
                t.entry_time.strftime("%Y-%m-%d %H:%M") if t.entry_time else None,
                t.exit_time.strftime("%Y-%m-%d %H:%M") if t.exit_time else None,
                t.notes,
            ]
            fill = win_fill if t.exit_result == "WIN" else loss_fill if t.exit_result == "LOSS" else None
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if fill:
                    cell.fill = fill

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
        ws.freeze_panes = "A2"

        # Stream the file
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        session.close()

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=trades_export.xlsx"}
        )
    finally:
        session.close()
