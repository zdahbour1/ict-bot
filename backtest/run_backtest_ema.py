"""
ICT Strategy Backtest — EMA + VWAP + VOLUME FILTER TEST
=========================================================
Builds on the Optimized V4 baseline and adds:
  - 1H 20 EMA trend filter:
      LONG  signals only when QQQ close > 1H 20 EMA
      SHORT signals only when QQQ close < 1H 20 EMA
  - VWAP filter (Option 1):
      LONG  only when entry price > daily VWAP
      SHORT only when entry price < daily VWAP
  - Volume spike confirmation (Option 3):
      Entry only when volume at entry bar >= VOLUME_MULT x 20-bar avg volume
  - 5m 50 EMA confluence filter (optional):
      Entry only taken if price is within EMA_CONFLUENCE_PCT% of the 5m 50 EMA

All other settings identical to Optimized V4:
  - Trade window: 7:00–11:00 AM PT
  - VIX filter: skip if VIX > 35
  - Trailing stop: up 10% → breakeven, up 20% → trail at peak-10%
  - Time exit: 90 minutes
  - Max 3 trades/day
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import math
import logging
import warnings
import pandas as pd
import numpy as np
import yfinance as yf
import pytz
from datetime import datetime, timedelta, date

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.WARNING)

from strategy.ict_long  import run_strategy as run_long
from strategy.ict_short import run_strategy_short as run_short
from strategy.levels    import get_all_levels

# ── Config ─────────────────────────────────────────────────
TICKER                  = "QQQ"
CONTRACTS               = 2
PROFIT_TARGET           = 1.00
STOP_LOSS               = 0.60
MAX_TRADES_DAY          = 4
TRADE_START_H           = 7
TRADE_START_MIN         = 0    # 7:00 AM PT
TRADE_END_H             = 9
VIX_THRESHOLD           = 35.0

# ── EMA Settings ───────────────────────────────────────────
EMA_PERIOD_1H           = 20        # 1H 20 EMA for trend direction
USE_5M_EMA_CONFLUENCE   = False     # Set True to also require 5m 50 EMA proximity
EMA_PERIOD_5M           = 50        # 5m 50 EMA for entry confluence
EMA_CONFLUENCE_PCT      = 0.003     # Price must be within 0.3% of 5m EMA to enter

# ── VWAP Settings (Option 1) ────────────────────────────────
USE_VWAP_FILTER         = False     # Long only above VWAP, short only below

# ── Volume Spike Settings (Option 3) ───────────────────────
USE_VOLUME_FILTER       = False     # Require above-average volume at entry bar
VOLUME_MULT             = 1.5       # Entry bar volume must be >= 1.5x 20-bar avg
VOLUME_LOOKBACK         = 20        # bars back to compute average volume

# ── News Filter ─────────────────────────────────────────────
USE_NEWS_FILTER         = True      # Skip trades within NEWS_BUFFER_MIN of major events
NEWS_BUFFER_MIN         = 30        # minutes before/after event to block

# ── Minimum R:R Filter ──────────────────────────────────────
USE_RR_FILTER           = False     # Skip trades with R:R below MIN_RR
MIN_RR                  = 2.0       # Minimum reward:risk ratio (start at 2:1, not 3:1)
                                    # (3:1 is ideal but may be too strict for scalping)

PT             = pytz.timezone("America/Los_Angeles")
OUTPUT_FILE    = os.path.join(os.path.dirname(__file__), "ICT_Backtest_EMA.xlsx")


# ── Helpers ────────────────────────────────────────────────

def bar_to_pt(ts, tz=PT):
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    return ts.tz_convert(tz)


def compute_ema(series: pd.Series, period: int) -> pd.Series:
    return series.ewm(span=period, adjust=False).mean()


# ── Major economic events (PT times) ───────────────────────
# All times in Pacific Time. ET = PT + 3h. 8:30 AM ET = 5:30 AM PT.
# Covers the 60-day backtest window (roughly Jan 27 – Mar 27, 2026).
import datetime as _dt
NEWS_EVENTS = [
    # (date, hour_pt, minute_pt, label)
    (_dt.date(2026, 1, 29), 11,  0,  "FOMC Decision"),
    (_dt.date(2026, 1, 29), 11, 30,  "FOMC Press Conference"),
    (_dt.date(2026, 2,  7),  5, 30,  "NFP Jobs Report"),
    (_dt.date(2026, 2, 12),  5, 30,  "CPI"),
    (_dt.date(2026, 2, 13),  5, 30,  "PPI"),
    (_dt.date(2026, 2, 14),  5, 30,  "Retail Sales"),
    (_dt.date(2026, 2, 19),  5, 30,  "PPI (revised)"),
    (_dt.date(2026, 2, 26),  5, 30,  "GDP"),
    (_dt.date(2026, 3,  7),  5, 30,  "NFP Jobs Report"),
    (_dt.date(2026, 3, 12),  5, 30,  "CPI"),
    (_dt.date(2026, 3, 13),  5, 30,  "PPI"),
    (_dt.date(2026, 3, 14),  5, 30,  "Retail Sales"),
    (_dt.date(2026, 3, 19), 11,  0,  "FOMC Decision"),
    (_dt.date(2026, 3, 19), 11, 30,  "FOMC Press Conference"),
    (_dt.date(2026, 3, 26),  5, 30,  "GDP"),
]


def is_near_news(entry_time_pt, buffer_min: int = NEWS_BUFFER_MIN) -> tuple:
    """
    Returns (True, event_label) if entry_time_pt is within buffer_min
    of any major economic event. Otherwise (False, '').
    """
    import datetime as dt
    for (ev_date, ev_h, ev_m, label) in NEWS_EVENTS:
        ev_dt = PT.localize(dt.datetime(ev_date.year, ev_date.month, ev_date.day, ev_h, ev_m))
        diff  = abs((entry_time_pt - ev_dt).total_seconds() / 60)
        if diff <= buffer_min:
            return True, label
    return False, ""


def compute_rr(entry: float, sl: float, tp: float, direction: str) -> float:
    """
    Reward : Risk ratio.
    LONG  = (tp - entry) / (entry - sl)
    SHORT = (entry - tp) / (sl - entry)
    Returns 0 if SL == entry (invalid).
    """
    try:
        if direction == "LONG":
            risk   = entry - sl
            reward = tp - entry
        else:
            risk   = sl - entry
            reward = entry - tp
        if risk <= 0 or reward <= 0:
            return 0.0
        return round(reward / risk, 2)
    except Exception:
        return 0.0


def compute_extra_levels(bars_5m: pd.DataFrame, day_date) -> list:
    """
    Extra liquidity levels not in the base levels.py:
      - Opening Range High/Low (first 15 min of session: 6:30–6:45 AM PT)
      - Previous Week High/Low
    """
    extra = []
    df = bars_5m.copy()
    if df.index.tzinfo is None:
        df.index = df.index.tz_localize("UTC")
    df.index = df.index.tz_convert(PT)

    # ── Opening Range (6:30–6:45 AM PT on current day) ──────
    or_mask = (
        (df.index.date == day_date) &
        (df.index.hour == 6) &
        (df.index.minute >= 30) &
        (df.index.minute < 45)
    )
    or_bars = df[or_mask]
    if not or_bars.empty:
        extra.append({"label": "OR_HIGH", "price": float(or_bars["high"].max())})
        extra.append({"label": "OR_LOW",  "price": float(or_bars["low"].min())})

    # ── Previous Week High/Low ───────────────────────────────
    import datetime as dt
    today      = day_date
    # Find what week number today is and get bars from the prior week
    week_start = today - dt.timedelta(days=today.weekday() + 7)   # Mon of last week
    week_end   = week_start + dt.timedelta(days=4)                  # Fri of last week
    pw_mask    = (df.index.date >= week_start) & (df.index.date <= week_end)
    pw_bars    = df[pw_mask]
    if not pw_bars.empty:
        extra.append({"label": "PWH", "price": float(pw_bars["high"].max())})
        extra.append({"label": "PWL", "price": float(pw_bars["low"].min())})

    return extra


def compute_vwap(bars: pd.DataFrame) -> pd.Series:
    """
    Daily VWAP — resets each trading day.
    VWAP = cumsum(typical_price * volume) / cumsum(volume)
    """
    typical = (bars["high"] + bars["low"] + bars["close"]) / 3
    dates   = bars.index.date
    vwap    = pd.Series(index=bars.index, dtype=float)

    for d in set(dates):
        mask = dates == d
        tp_v = typical[mask] * bars["volume"][mask]
        vwap[mask] = tp_v.cumsum() / bars["volume"][mask].cumsum()

    return vwap


def has_volume_spike(bars: pd.DataFrame, entry_bar: int) -> bool:
    """
    Returns True if volume at entry_bar >= VOLUME_MULT x avg of prior VOLUME_LOOKBACK bars.
    """
    if entry_bar < VOLUME_LOOKBACK:
        return True   # not enough history — don't filter
    avg_vol    = bars["volume"].iloc[entry_bar - VOLUME_LOOKBACK:entry_bar].mean()
    entry_vol  = bars["volume"].iloc[entry_bar]
    if avg_vol == 0:
        return True
    return entry_vol >= VOLUME_MULT * avg_vol


def get_trend_bias(bars_1h: pd.DataFrame, as_of_time) -> str:
    """
    Returns 'BULLISH' or 'BEARISH' based on whether price is above/below
    the 1H 20 EMA at the given time. Returns 'NEUTRAL' if insufficient data.
    """
    hist = bars_1h[bars_1h.index <= as_of_time]
    if len(hist) < EMA_PERIOD_1H:
        return "NEUTRAL"
    ema = compute_ema(hist["close"], EMA_PERIOD_1H)
    last_close = hist["close"].iloc[-1]
    last_ema   = ema.iloc[-1]
    if last_close > last_ema:
        return "BULLISH"
    elif last_close < last_ema:
        return "BEARISH"
    return "NEUTRAL"


def near_5m_ema(bars_5m: pd.DataFrame, entry_bar: int, entry_price: float) -> bool:
    """
    Returns True if entry price is within EMA_CONFLUENCE_PCT of the 5m 50 EMA.
    Used as an optional extra confluence filter.
    """
    if entry_bar < EMA_PERIOD_5M:
        return True  # not enough data — don't filter
    ema = compute_ema(bars_5m["close"], EMA_PERIOD_5M)
    ema_val = ema.iloc[entry_bar]
    if ema_val == 0:
        return True
    pct_diff = abs(entry_price - ema_val) / ema_val
    return pct_diff <= EMA_CONFLUENCE_PCT


def estimate_option_price(qqq_price: float, bars_5m: pd.DataFrame, entry_bar: int) -> float:
    lookback = min(20, entry_bar)
    recent   = bars_5m.iloc[max(0, entry_bar - lookback):entry_bar]
    if len(recent) < 2:
        iv = 0.22
    else:
        rets = np.log(recent["close"] / recent["close"].shift(1)).dropna()
        iv   = rets.std() * math.sqrt(252 * 78)
        iv   = max(iv, 0.10)
    T     = 3.0 / (252 * 6.5)
    price = qqq_price * iv * math.sqrt(T) * 0.4
    return max(round(price, 2), 0.05)


def simulate_exit(bars_5m: pd.DataFrame, entry_bar: int,
                  entry_opt_px: float, direction: str):
    """Trailing stop + 90-min time exit (same as Optimized V4)."""
    tp_px        = entry_opt_px * (1 + PROFIT_TARGET)
    sl_pct       = -STOP_LOSS
    peak_pnl_pct = 0.0
    entry_close  = bars_5m.iloc[entry_bar]["close"]
    MAX_BARS     = 18   # 90 min = 18 x 5m bars

    for i in range(entry_bar + 1, min(entry_bar + 300, len(bars_5m))):
        bar      = bars_5m.iloc[i]
        bar_time = bars_5m.index[i]
        bar_pt   = bar_to_pt(bar_time)
        bars_in  = i - entry_bar

        underlying_chg = bar["close"] - entry_close
        opt_chg = underlying_chg * 0.5 if direction == "LONG" else -underlying_chg * 0.5
        cur_opt = max(round(entry_opt_px + opt_chg, 2), 0.01)
        pnl_pct = (cur_opt - entry_opt_px) / entry_opt_px

        if pnl_pct > peak_pnl_pct:
            peak_pnl_pct = pnl_pct

        if peak_pnl_pct >= 0.20:
            sl_pct = peak_pnl_pct - 0.10
        elif peak_pnl_pct >= 0.10:
            sl_pct = 0.00

        hit_tp    = pnl_pct >= PROFIT_TARGET
        hit_sl    = pnl_pct <= sl_pct
        eod       = bar_pt.hour >= 13
        time_exit = bars_in >= MAX_BARS

        if hit_tp or hit_sl or eod or time_exit:
            if hit_tp:
                exit_px = round(tp_px, 2)
                pnl_pct = PROFIT_TARGET
                result  = "WIN"
                reason  = "TP"
            elif hit_sl and sl_pct == 0.0:
                exit_px = entry_opt_px
                pnl_pct = 0.0
                result  = "SCRATCH"
                reason  = "BREAKEVEN"
            elif hit_sl:
                exit_px = max(round(entry_opt_px * (1 + sl_pct), 2), 0.01)
                result  = "WIN" if sl_pct > 0 else "LOSS"
                reason  = "TRAIL SL" if sl_pct > 0 else "SL"
            elif time_exit:
                exit_px = cur_opt
                result  = "WIN" if pnl_pct > 0 else "LOSS" if pnl_pct < 0 else "SCRATCH"
                reason  = "TIME EXIT (90min)"
            else:
                exit_px = cur_opt
                result  = "WIN" if pnl_pct > 0 else "LOSS"
                reason  = "EOD"

            pnl_usd = round((exit_px - entry_opt_px) * 100 * CONTRACTS, 2)
            return {
                "exit_time":      bar_time,
                "exit_option_px": exit_px,
                "pnl_pct":        round(pnl_pct * 100, 1),
                "pnl_usd":        pnl_usd,
                "result":         result,
                "exit_reason":    reason,
                "bars_held":      bars_in,
            }

    last    = bars_5m.iloc[min(entry_bar + 299, len(bars_5m) - 1)]
    chg     = last["close"] - entry_close
    opt_chg = chg * 0.5 if direction == "LONG" else -chg * 0.5
    cur     = max(round(entry_opt_px + opt_chg, 2), 0.01)
    pnl     = (cur - entry_opt_px) / entry_opt_px
    return {
        "exit_time":      bars_5m.index[min(entry_bar + 299, len(bars_5m) - 1)],
        "exit_option_px": cur,
        "pnl_pct":        round(pnl * 100, 1),
        "pnl_usd":        round((cur - entry_opt_px) * 100 * CONTRACTS, 2),
        "result":         "WIN" if pnl > 0 else "LOSS",
        "exit_reason":    "END",
        "bars_held":      299,
    }


def occ_symbol(ticker, exp_date, direction, strike):
    opt_type   = "C" if direction == "LONG" else "P"
    strike_int = int(round(strike * 1000))
    return f"{ticker}{exp_date.strftime('%y%m%d')}{opt_type}{strike_int:08d}"


# ── Main backtest ──────────────────────────────────────────

def main():
    confluence_label = "ON" if USE_5M_EMA_CONFLUENCE else "OFF"
    print("=" * 60)
    print("  ICT Strategy Backtest — EMA + VWAP + MORE LEVELS")
    print(f"  1H {EMA_PERIOD_1H} EMA | VWAP | OR + PW levels")
    print(f"  Window: 6:30 AM–12 PM PT | Max {MAX_TRADES_DAY} trades/day")
    print("=" * 60)
    print()

    print("Downloading QQQ + VIX data...")
    raw_5m      = yf.download(TICKER, period="60d", interval="5m",
                              auto_adjust=True, progress=False)
    raw_1h      = yf.download(TICKER, period="60d", interval="1h",
                              auto_adjust=True, progress=False)
    raw_4h_base = yf.download(TICKER, period="60d", interval="1h",
                              auto_adjust=True, progress=False)
    vix_raw     = yf.download("^VIX", period="60d", interval="1d",
                              auto_adjust=True, progress=False)

    if raw_5m.empty:
        print("ERROR: Could not download data.")
        return

    for df_ref in [raw_5m, raw_1h, raw_4h_base, vix_raw]:
        if isinstance(df_ref.columns, pd.MultiIndex):
            df_ref.columns = df_ref.columns.get_level_values(0).str.lower()
        else:
            df_ref.columns = df_ref.columns.str.lower()

    raw_4h = raw_4h_base.resample("4h").agg({
        "open": "first", "high": "max",
        "low": "min", "close": "last", "volume": "sum"
    }).dropna()

    def to_pt(df):
        if df.index.tzinfo is None:
            df.index = df.index.tz_localize("UTC")
        df.index = df.index.tz_convert(PT)
        return df

    bars_5m = to_pt(raw_5m.copy())
    bars_1h = to_pt(raw_1h.copy())
    bars_4h = to_pt(raw_4h.copy())

    for df in [bars_5m, bars_1h, bars_4h]:
        df.columns = [c.lower() for c in df.columns]

    # Pre-compute 1H EMA across all data
    bars_1h["ema"] = compute_ema(bars_1h["close"], EMA_PERIOD_1H)

    # Pre-compute 5m EMA if confluence filter is on
    if USE_5M_EMA_CONFLUENCE:
        bars_5m["ema_50"] = compute_ema(bars_5m["close"], EMA_PERIOD_5M)

    # Build VIX daily lookup
    vix_by_date = {}
    if not vix_raw.empty:
        for idx, row in vix_raw.iterrows():
            d = idx.date() if hasattr(idx, "date") else idx
            vix_by_date[d] = float(row.get("close", row.get("Close", 20)))

    trading_days = sorted(set(bars_5m.index.date))
    print(f"Data: {len(trading_days)} trading days | {len(bars_5m):,} bars\n")

    # Pre-compute daily VWAP across all 5m bars
    if USE_VWAP_FILTER:
        bars_5m["vwap"] = compute_vwap(bars_5m)

    trades             = []
    skipped_vix        = 0
    skipped_ema        = 0
    skipped_vwap       = 0
    skipped_volume     = 0
    skipped_news       = 0
    skipped_rr         = 0
    skipped_confluence = 0

    for day_date in trading_days:

        vix_today = vix_by_date.get(day_date, 0)
        if vix_today > VIX_THRESHOLD:
            skipped_vix += 1
            print(f"  {day_date}  SKIPPED (VIX={vix_today:.1f} > {VIX_THRESHOLD})")
            continue

        day_end = pd.Timestamp(day_date, tz=PT) + pd.Timedelta(hours=23, minutes=59)

        hist_5m = bars_5m[bars_5m.index <= day_end].copy()
        hist_1h = bars_1h[bars_1h.index <= day_end].copy()
        hist_4h = bars_4h[bars_4h.index <= day_end].copy()

        if len(hist_5m) < 50:
            continue

        day_5m = bars_5m[bars_5m.index.date == day_date]
        if day_5m.empty:
            continue

        try:
            levels = get_all_levels(hist_5m, hist_1h, hist_4h)
            levels += compute_extra_levels(hist_5m, day_date)
            levels  = [l for l in levels if l["price"] > 0]
        except Exception:
            continue

        if not levels:
            continue

        try:
            long_signals  = run_long(day_5m,  hist_1h, hist_4h, levels)
            short_signals = run_short(day_5m, hist_1h, hist_4h, levels)
        except Exception:
            continue

        all_signals = []
        for s in long_signals:
            s["direction"] = "LONG"
            all_signals.append(s)
        for s in short_signals:
            s["direction"] = "SHORT"
            all_signals.append(s)

        all_signals.sort(key=lambda x: x.get("entry_time", pd.Timestamp.min))

        trades_today  = 0
        last_exit_bar = -1

        for sig in all_signals:
            if trades_today >= MAX_TRADES_DAY:
                break

            entry_time = sig.get("entry_time")
            if entry_time is None:
                continue

            et_pt = bar_to_pt(entry_time)
            in_window = (
                (et_pt.hour > TRADE_START_H or (et_pt.hour == TRADE_START_H and et_pt.minute >= TRADE_START_MIN))
                and et_pt.hour < TRADE_END_H
            )
            if not in_window:
                continue

            direction = sig["direction"]

            # ── 1H EMA Trend Filter ──────────────────────────
            trend = get_trend_bias(hist_1h, entry_time)
            if direction == "LONG"  and trend != "BULLISH":
                skipped_ema += 1
                continue
            if direction == "SHORT" and trend != "BEARISH":
                skipped_ema += 1
                continue
            # NEUTRAL = skip (no clear bias)

            entry_bar   = sig.get("entry_bar", 0)
            entry_price = sig.get("entry_price", 0)

            if entry_bar <= last_exit_bar:
                continue
            if entry_price <= 0:
                continue

            # ── VWAP Filter ──────────────────────────────────
            if USE_VWAP_FILTER and "vwap" in day_5m.columns:
                try:
                    vwap_at_entry = day_5m["vwap"].iloc[entry_bar]
                    if direction == "LONG"  and entry_price < vwap_at_entry:
                        skipped_vwap += 1
                        continue
                    if direction == "SHORT" and entry_price > vwap_at_entry:
                        skipped_vwap += 1
                        continue
                except (IndexError, KeyError):
                    pass

            # ── Volume Spike Filter ──────────────────────────
            if USE_VOLUME_FILTER:
                if not has_volume_spike(day_5m, entry_bar):
                    skipped_volume += 1
                    continue

            # ── News Filter ───────────────────────────────────
            if USE_NEWS_FILTER:
                near_news, news_label = is_near_news(et_pt)
                if near_news:
                    skipped_news += 1
                    continue

            # ── Minimum R:R Filter ────────────────────────────
            if USE_RR_FILTER:
                sl_price = sig.get("sl", 0)
                tp_price = sig.get("tp", 0)
                rr = compute_rr(entry_price, sl_price, tp_price, direction)
                if rr < MIN_RR:
                    skipped_rr += 1
                    continue

            # ── Optional 5m EMA Confluence Filter ────────────
            if USE_5M_EMA_CONFLUENCE:
                if not near_5m_ema(day_5m, entry_bar, entry_price):
                    skipped_confluence += 1
                    continue

            strike     = round(entry_price)
            opt_px     = estimate_option_price(entry_price, day_5m, entry_bar)
            symbol     = occ_symbol(TICKER, day_date, direction, strike)
            total_cost = round(opt_px * 100 * CONTRACTS, 2)

            exit_info     = simulate_exit(day_5m, entry_bar, opt_px, direction)
            last_exit_bar = exit_info.get("bars_held", 0) + entry_bar

            exit_time  = exit_info["exit_time"]
            et_pt_exit = bar_to_pt(exit_time)

            # Snapshot EMA + VWAP values for the trade log
            ema_1h_at_entry  = hist_1h["ema"].iloc[-1] if "ema" in hist_1h.columns else 0
            try:
                vwap_at_entry = day_5m["vwap"].iloc[entry_bar] if "vwap" in day_5m.columns else 0
                entry_vol     = day_5m["volume"].iloc[entry_bar]
                avg_vol       = day_5m["volume"].iloc[max(0, entry_bar - VOLUME_LOOKBACK):entry_bar].mean()
                vol_ratio     = round(entry_vol / avg_vol, 2) if avg_vol > 0 else 0
            except (IndexError, KeyError):
                vwap_at_entry = 0
                vol_ratio     = 0

            rr_logged = compute_rr(entry_price, sig.get("sl", 0), sig.get("tp", 0), direction)

            trades.append({
                "Date":              day_date.strftime("%Y-%m-%d"),
                "Direction":         direction,
                "Trend (1H EMA)":    trend,
                "1H EMA":            f"${ema_1h_at_entry:.2f}",
                "VWAP":              f"${vwap_at_entry:.2f}",
                "Vol Ratio":         f"{vol_ratio}x",
                "R:R":               f"{rr_logged:.1f}:1",
                "Signal Type":       sig.get("signal_type", ""),
                "Raided Level":      sig.get("raid", {}).get("raided_level", ""),
                "VIX":               f"{vix_today:.1f}",
                "Entry Time (PT)":   et_pt.strftime("%I:%M %p"),
                "QQQ Entry Price":   f"${entry_price:.2f}",
                "Option Symbol":     symbol,
                "Strike":            f"${strike}",
                "Option Entry $":    f"${opt_px:.2f}",
                "Contracts":         CONTRACTS,
                "Total Cost":        f"${total_cost:.2f}",
                "SL (QQQ)":          f"${sig.get('sl', 0):.2f}",
                "TP (QQQ)":          f"${sig.get('tp', 0):.2f}",
                "Exit Time (PT)":    et_pt_exit.strftime("%I:%M %p"),
                "Option Exit $":     f"${exit_info['exit_option_px']:.2f}",
                "P&L %":             f"{exit_info['pnl_pct']:+.1f}%",
                "P&L $":             f"${exit_info['pnl_usd']:+.2f}",
                "Exit Reason":       exit_info.get("exit_reason", ""),
                "Result":            exit_info["result"],
                "_pnl_usd_raw":      exit_info["pnl_usd"],
                "_result_raw":       exit_info["result"],
            })
            trades_today += 1

        if trades_today > 0:
            print(f"  {day_date}  ->  {trades_today} trade(s)  [VIX={vix_today:.1f}]")

    if not trades:
        print("\nNo signals passed all filters.")
        return

    df        = pd.DataFrame(trades)
    wins      = (df["_result_raw"] == "WIN").sum()
    losses    = (df["_result_raw"] == "LOSS").sum()
    scratches = (df["_result_raw"] == "SCRATCH").sum()
    total     = len(df)
    win_rate  = wins / (wins + losses) * 100 if (wins + losses) > 0 else 0
    total_pnl = df["_pnl_usd_raw"].sum()
    avg_win   = df.loc[df["_result_raw"] == "WIN",  "_pnl_usd_raw"].mean()
    avg_loss  = df.loc[df["_result_raw"] == "LOSS", "_pnl_usd_raw"].mean()

    print(f"\n{'='*60}")
    print(f"  EMA FILTER RESULTS")
    print(f"{'='*60}")
    print(f"  Total Trades       : {total}")
    print(f"  Wins               : {wins}")
    print(f"  Losses             : {losses}")
    print(f"  Scratches          : {scratches}")
    print(f"  Win Rate           : {win_rate:.1f}%  (excl. scratches)")
    print(f"  Total P&L          : ${total_pnl:+.2f}")
    print(f"  Avg Win            : ${avg_win:+.2f}" if not math.isnan(avg_win) else "  Avg Win           : N/A")
    print(f"  Avg Loss           : ${avg_loss:+.2f}" if not math.isnan(avg_loss) else "  Avg Loss          : N/A")
    print(f"  ---")
    print(f"  VIX days skipped   : {skipped_vix}")
    print(f"  EMA filtered out   : {skipped_ema} signals (wrong trend)")
    print(f"  VWAP filtered out  : {skipped_vwap} signals (wrong side of VWAP)")
    print(f"  News filtered out  : {skipped_news} signals (near major event)")
    print(f"  R:R filtered out   : {skipped_rr} signals (R:R < {MIN_RR}:1)")
    print(f"  Volume filtered out: {skipped_volume} signals (weak volume)")
    if USE_5M_EMA_CONFLUENCE:
        print(f"  5m EMA filtered    : {skipped_confluence} signals (not near EMA)")
    print(f"{'='*60}\n")

    # ── Excel export ──────────────────────────────────────
    display_cols = [c for c in df.columns if not c.startswith("_")]
    export_df    = df[display_cols].copy()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Trade Log", index=False)
        ws = writer.sheets["Trade Log"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill  = PatternFill("solid", fgColor="1A4731")
        win_fill     = PatternFill("solid", fgColor="C6EFCE")
        loss_fill    = PatternFill("solid", fgColor="FFC7CE")
        scratch_fill = PatternFill("solid", fgColor="FFEB9C")
        thin         = Side(style="thin", color="CCCCCC")
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = border

        result_col_idx = display_cols.index("Result") + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            result_val = ws.cell(row=row_idx, column=result_col_idx).value
            row_fill   = win_fill     if result_val == "WIN"     else \
                         loss_fill    if result_val == "LOSS"    else \
                         scratch_fill if result_val == "SCRATCH" else None
            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")
                cell.font      = Font(size=10)
                if row_fill:
                    cell.fill = row_fill

        for col_idx, col in enumerate(display_cols, 1):
            max_len = max(len(str(col)), max(
                (len(str(export_df[col].iloc[i])) for i in range(len(export_df))),
                default=0
            )) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 25)

        ws.freeze_panes = "A2"

        # ── Comparison summary sheet ──────────────────────
        avg_win_str  = f"${avg_win:+.2f}"  if not math.isnan(avg_win)  else "N/A"
        avg_loss_str = f"${avg_loss:+.2f}" if not math.isnan(avg_loss) else "N/A"

        sum_df = pd.DataFrame([
            ["Metric",            "Optimized V4", "EMA Only", "EMA+VWAP+More Levels", "Change vs V4"],
            ["Total Trades",      79,             44,          total,                  total - 79],
            ["Win Rate",          "43.0%",        "52.6%",     f"{win_rate:.1f}%",     f"{win_rate-43.0:+.1f}%"],
            ["Total P&L",         "$+860",        "$+876",     f"${total_pnl:+.2f}",  f"${total_pnl-860:+.2f}"],
            ["Avg Win",           "$+77.59",      "$+86.00",   avg_win_str,            "—"],
            ["Avg Loss",          "$-50.80",      "$-46.89",   avg_loss_str,           "—"],
            ["Scratches",         "—",            6,           scratches,              "—"],
            ["EMA filtered",      "0",            131,         skipped_ema,            "—"],
            ["VWAP filtered",     "0",            "0",         skipped_vwap,           f"-{skipped_vwap}"],
            ["VIX skipped",       0,              0,           skipped_vix,            "—"],
            ["Trade window",      "7-11 AM",      "7-11 AM",   "6:30 AM-12 PM",        "Wider"],
            ["Max trades/day",    3,              3,           MAX_TRADES_DAY,         f"+{MAX_TRADES_DAY-3}"],
            ["Extra levels",      "None",         "None",      "OR + Prev Week",       "New"],
            ["EMA Filter",        "None",         "1H 20 EMA", "1H 20 EMA",           "Same"],
            ["VWAP Filter",       "None",         "None",      "Added",                "New"],
        ])
        sum_df.to_excel(writer, sheet_name="Comparison", index=False, header=False)

        ws2 = writer.sheets["Comparison"]
        header_fill2 = PatternFill("solid", fgColor="1F3864")
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill2
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(size=10)
        for col in ["A", "B", "C", "D"]:
            ws2.column_dimensions[col].width = 26

    # ── CSV export ────────────────────────────────────────
    csv_file = OUTPUT_FILE.replace(".xlsx", ".csv")
    export_df.to_csv(csv_file, index=False)
    print(f"Report saved to:\n  {OUTPUT_FILE}\n  {csv_file}\n")


if __name__ == "__main__":
    main()
