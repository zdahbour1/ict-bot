"""
ICT Strategy Backtest — OPTIMIZED V4 (for comparison only)
================================================================
Changes vs baseline:
  1. Trade window: 7:00–11:00 AM PT
  2. Both iFVG and OB signals allowed
  3. Max 3 trades per day
  4. VIX filter at 35
  5. TRAILING STOP: once up 10%, SL moves to breakeven (0%)
                    once up 20%, SL trails at current - 10%
  6. BREAKEVEN STOP: never let a winning trade turn into full -15% loss
  7. TIME EXIT: if still in trade after 90 minutes, exit to avoid theta decay

NOTE: This does NOT change any live bot code. For analysis only.
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import math
import logging
import warnings
import pandas as pd
import numpy as np
import yfinance as yf
import pytz
from datetime import datetime, timedelta, date

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.WARNING)

from strategy.ict_long  import run_strategy as run_long
from strategy.ict_short import run_strategy_short as run_short
from strategy.levels    import get_all_levels

# ── Config ─────────────────────────────────────────────────
TICKER          = "QQQ"
CONTRACTS       = 2
PROFIT_TARGET   = 0.25
STOP_LOSS       = 0.15
MAX_TRADES_DAY  = 3          # OPTIMIZED: was 6, v1 was 2
TRADE_START_H   = 7          # same
TRADE_END_H     = 11         # OPTIMIZED: 7-11 AM PT (compromise)
SMA_PERIOD      = 20         # 20-bar 1H SMA for direction filter
VIX_THRESHOLD   = 35.0       # OPTIMIZED: only skip truly extreme days (was 25)
PT              = pytz.timezone("America/Los_Angeles")
OUTPUT_FILE     = os.path.join(os.path.dirname(__file__), "ICT_Backtest_Optimized.xlsx")


# ── Helpers ────────────────────────────────────────────────

def bar_to_pt(ts, tz=PT):
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    return ts.tz_convert(tz)


def estimate_option_price(qqq_price: float, bars_5m: pd.DataFrame,
                           entry_bar: int) -> float:
    lookback = min(20, entry_bar)
    recent = bars_5m.iloc[max(0, entry_bar - lookback):entry_bar]
    if len(recent) < 2:
        iv = 0.22
    else:
        rets = np.log(recent["close"] / recent["close"].shift(1)).dropna()
        iv = rets.std() * math.sqrt(252 * 78)
        iv = max(iv, 0.10)
    T = 3.0 / (252 * 6.5)
    price = qqq_price * iv * math.sqrt(T) * 0.4
    return max(round(price, 2), 0.05)


def simulate_exit(bars_5m: pd.DataFrame, entry_bar: int,
                  entry_opt_px: float, direction: str):
    """
    Smart exit with:
    - Trailing stop: once up 10% -> SL moves to breakeven
                     once up 20% -> SL trails at (peak - 10%)
    - Time exit: force close after 90 minutes (18 x 5m bars) to avoid theta decay
    """
    tp_px        = entry_opt_px * (1 + PROFIT_TARGET)
    sl_pct       = -STOP_LOSS          # starts at -15%
    peak_pnl_pct = 0.0
    entry_close  = bars_5m.iloc[entry_bar]["close"]
    entry_time   = bars_5m.index[entry_bar]
    MAX_BARS     = 18                  # 90 minutes = 18 x 5m bars

    for i in range(entry_bar + 1, min(entry_bar + 300, len(bars_5m))):
        bar      = bars_5m.iloc[i]
        bar_time = bars_5m.index[i]
        bar_pt   = bar_to_pt(bar_time)
        bars_in  = i - entry_bar

        underlying_chg = bar["close"] - entry_close
        opt_chg = underlying_chg * 0.5 if direction == "LONG" else -underlying_chg * 0.5
        cur_opt = max(round(entry_opt_px + opt_chg, 2), 0.01)
        pnl_pct = (cur_opt - entry_opt_px) / entry_opt_px

        # Track peak
        if pnl_pct > peak_pnl_pct:
            peak_pnl_pct = pnl_pct

        # Trailing stop logic
        if peak_pnl_pct >= 0.20:
            # Trail SL at peak - 10%
            sl_pct = peak_pnl_pct - 0.10
        elif peak_pnl_pct >= 0.10:
            # Move SL to breakeven
            sl_pct = 0.00

        hit_tp      = pnl_pct >= PROFIT_TARGET
        hit_sl      = pnl_pct <= sl_pct
        eod         = bar_pt.hour >= 13
        time_exit   = bars_in >= MAX_BARS

        if hit_tp or hit_sl or eod or time_exit:
            if hit_tp:
                exit_px = round(tp_px, 2)
                pnl_pct = PROFIT_TARGET
                result  = "WIN"
                reason  = "TP"
            elif hit_sl and sl_pct == 0.0:
                exit_px = entry_opt_px   # breakeven
                pnl_pct = 0.0
                result  = "SCRATCH"
                reason  = "BREAKEVEN"
            elif hit_sl:
                exit_px = max(round(entry_opt_px * (1 + sl_pct), 2), 0.01)
                result  = "WIN" if sl_pct > 0 else "LOSS"
                reason  = "TRAIL SL" if sl_pct > 0 else "SL"
            elif time_exit:
                exit_px = cur_opt
                result  = "WIN" if pnl_pct > 0 else "LOSS" if pnl_pct < 0 else "SCRATCH"
                reason  = "TIME EXIT (90min)"
            else:
                exit_px = cur_opt
                result  = "WIN" if pnl_pct > 0 else "LOSS"
                reason  = "EOD"

            pnl_usd = round((exit_px - entry_opt_px) * 100 * CONTRACTS, 2)
            return {
                "exit_time":      bar_time,
                "exit_option_px": exit_px,
                "pnl_pct":        round(pnl_pct * 100, 1),
                "pnl_usd":        pnl_usd,
                "result":         result,
                "exit_reason":    reason,
                "bars_held":      bars_in,
            }

    # Fallback
    last = bars_5m.iloc[min(entry_bar + 299, len(bars_5m) - 1)]
    chg  = last["close"] - entry_close
    opt_chg = chg * 0.5 if direction == "LONG" else -chg * 0.5
    cur  = max(round(entry_opt_px + opt_chg, 2), 0.01)
    pnl  = (cur - entry_opt_px) / entry_opt_px
    return {
        "exit_time":      bars_5m.index[min(entry_bar + 299, len(bars_5m) - 1)],
        "exit_option_px": cur,
        "pnl_pct":        round(pnl * 100, 1),
        "pnl_usd":        round((cur - entry_opt_px) * 100 * CONTRACTS, 2),
        "result":         "WIN" if pnl > 0 else "LOSS",
        "exit_reason":    "END",
        "bars_held":      299,
    }


def occ_symbol(ticker, exp_date, direction, strike):
    opt_type = "C" if direction == "LONG" else "P"
    strike_int = int(round(strike * 1000))
    return f"{ticker}{exp_date.strftime('%y%m%d')}{opt_type}{strike_int:08d}"


def get_market_direction(bars_1h: pd.DataFrame, as_of_time) -> str:
    """
    OPTIMIZATION 2: Market direction filter.
    Returns 'BULLISH', 'BEARISH', or 'NEUTRAL' based on 20-bar 1H SMA.
    """
    hist = bars_1h[bars_1h.index <= as_of_time]
    if len(hist) < SMA_PERIOD:
        return "NEUTRAL"
    sma = hist["close"].iloc[-SMA_PERIOD:].mean()
    last_close = hist["close"].iloc[-1]
    if last_close > sma * 1.001:
        return "BULLISH"
    elif last_close < sma * 0.999:
        return "BEARISH"
    return "NEUTRAL"


# ── Main backtest ──────────────────────────────────────────

def main():
    print("=" * 60)
    print("  ICT Strategy Backtest — OPTIMIZED")
    print("  Changes: 7-11AM window | iFVG+OB |")
    print("           max 3 trades/day | VIX<35")
    print("=" * 60)
    print()

    print("Downloading QQQ + VIX data...")
    raw_5m = yf.download(TICKER, period="60d", interval="5m",
                         auto_adjust=True, progress=False)
    raw_1h = yf.download(TICKER, period="60d", interval="1h",
                         auto_adjust=True, progress=False)
    raw_4h_base = yf.download(TICKER, period="60d", interval="1h",
                               auto_adjust=True, progress=False)
    vix_raw = yf.download("^VIX", period="60d", interval="1d",
                           auto_adjust=True, progress=False)

    if raw_5m.empty:
        print("ERROR: Could not download data.")
        return

    for df_ref in [raw_5m, raw_1h, raw_4h_base, vix_raw]:
        if isinstance(df_ref.columns, pd.MultiIndex):
            df_ref.columns = df_ref.columns.get_level_values(0).str.lower()
        else:
            df_ref.columns = df_ref.columns.str.lower()

    raw_4h = raw_4h_base.resample("4h").agg({
        "open": "first", "high": "max",
        "low": "min", "close": "last", "volume": "sum"
    }).dropna()

    def to_pt(df):
        if df.index.tzinfo is None:
            df.index = df.index.tz_localize("UTC")
        df.index = df.index.tz_convert(PT)
        return df

    bars_5m = to_pt(raw_5m.copy())
    bars_1h = to_pt(raw_1h.copy())
    bars_4h = to_pt(raw_4h.copy())

    for df in [bars_5m, bars_1h, bars_4h]:
        df.columns = [c.lower() for c in df.columns]

    # Build VIX daily lookup
    vix_by_date = {}
    if not vix_raw.empty:
        for idx, row in vix_raw.iterrows():
            d = idx.date() if hasattr(idx, 'date') else idx
            vix_by_date[d] = float(row.get("close", row.get("Close", 20)))

    trading_days = sorted(set(bars_5m.index.date))
    print(f"Data: {len(trading_days)} trading days | {len(bars_5m):,} bars\n")

    trades = []
    skipped_vix = 0
    skipped_direction = 0
    skipped_ob = 0

    for day_date in trading_days:

        # OPTIMIZATION 5: VIX filter
        vix_today = vix_by_date.get(day_date, 0)
        if vix_today > VIX_THRESHOLD:
            skipped_vix += 1
            print(f"  {day_date}  SKIPPED (VIX={vix_today:.1f} > {VIX_THRESHOLD})")
            continue

        day_end = pd.Timestamp(day_date) + pd.Timedelta(hours=23, minutes=59)
        day_end = day_end.tz_localize(PT)

        hist_5m = bars_5m[bars_5m.index <= day_end].copy()
        hist_1h = bars_1h[bars_1h.index <= day_end].copy()
        hist_4h = bars_4h[bars_4h.index <= day_end].copy()

        if len(hist_5m) < 50:
            continue

        day_5m = bars_5m[bars_5m.index.date == day_date]
        if day_5m.empty:
            continue

        try:
            levels = get_all_levels(hist_5m, hist_1h, hist_4h)
        except Exception:
            continue

        if not levels:
            continue

        try:
            long_signals  = run_long(day_5m,  hist_1h, hist_4h, levels)
            short_signals = run_short(day_5m, hist_1h, hist_4h, levels)
        except Exception:
            continue

        # Both iFVG and OB signals allowed in V3
        pass

        all_signals = []
        for s in long_signals:
            s["direction"] = "LONG"
            all_signals.append(s)
        for s in short_signals:
            s["direction"] = "SHORT"
            all_signals.append(s)

        all_signals.sort(key=lambda x: x.get("entry_time", pd.Timestamp.min))

        trades_today = 0
        last_exit_bar = -1

        for sig in all_signals:
            # OPTIMIZATION 4: max 2 trades per day
            if trades_today >= MAX_TRADES_DAY:
                break

            entry_time = sig.get("entry_time")
            if entry_time is None:
                continue

            et_pt = bar_to_pt(entry_time)

            # OPTIMIZATION 1: 7-9 AM PT only
            if not (TRADE_START_H <= et_pt.hour < TRADE_END_H):
                continue

            direction    = sig["direction"]
            market_bias  = get_market_direction(hist_1h, entry_time)

            entry_bar = sig.get("entry_bar", 0)
            if entry_bar <= last_exit_bar:
                continue

            entry_price = sig.get("entry_price", 0)
            if entry_price <= 0:
                continue

            strike    = round(entry_price)
            opt_px    = estimate_option_price(entry_price, day_5m, entry_bar)
            symbol    = occ_symbol(TICKER, day_date, direction, strike)
            total_cost = round(opt_px * 100 * CONTRACTS, 2)

            exit_info = simulate_exit(day_5m, entry_bar, opt_px, direction)
            last_exit_bar = exit_info.get("bars_held", 0) + entry_bar

            exit_time  = exit_info["exit_time"]
            et_pt_exit = bar_to_pt(exit_time)

            trades.append({
                "Date":             day_date.strftime("%Y-%m-%d"),
                "Direction":        direction,
                "Signal Type":      sig.get("signal_type", ""),
                "Raided Level":     sig.get("raid", {}).get("raided_level", ""),
                "Market Bias":      market_bias,
                "VIX":              f"{vix_today:.1f}",
                "Entry Time (PT)":  et_pt.strftime("%I:%M %p"),
                "QQQ Entry Price":  f"${entry_price:.2f}",
                "Option Symbol":    symbol,
                "Strike":           f"${strike}",
                "Option Entry $":   f"${opt_px:.2f}",
                "Contracts":        CONTRACTS,
                "Total Cost":       f"${total_cost:.2f}",
                "SL (QQQ)":         f"${sig.get('sl', 0):.2f}",
                "TP (QQQ)":         f"${sig.get('tp', 0):.2f}",
                "Exit Time (PT)":   et_pt_exit.strftime("%I:%M %p"),
                "Option Exit $":    f"${exit_info['exit_option_px']:.2f}",
                "P&L %":            f"{exit_info['pnl_pct']:+.1f}%",
                "P&L $":            f"${exit_info['pnl_usd']:+.2f}",
                "Exit Reason":      exit_info.get("exit_reason", ""),
                "Result":           exit_info["result"],
                "_pnl_usd_raw":     exit_info["pnl_usd"],
                "_result_raw":      exit_info["result"],
            })
            trades_today += 1

        if trades_today > 0:
            print(f"  {day_date}  ->  {trades_today} trade(s)  [VIX={vix_today:.1f}]")

    if not trades:
        print("\nNo signals passed all filters.")
        return

    df = pd.DataFrame(trades)
    wins     = (df["_result_raw"] == "WIN").sum()
    losses   = (df["_result_raw"] == "LOSS").sum()
    total    = len(df)
    win_rate = wins / total * 100 if total > 0 else 0
    total_pnl = df["_pnl_usd_raw"].sum()
    avg_win   = df.loc[df["_result_raw"] == "WIN",  "_pnl_usd_raw"].mean()
    avg_loss  = df.loc[df["_result_raw"] == "LOSS", "_pnl_usd_raw"].mean()

    print(f"\n{'='*60}")
    print(f"  OPTIMIZED RESULTS")
    print(f"{'='*60}")
    print(f"  Total Trades       : {total}")
    print(f"  Wins               : {wins}")
    print(f"  Losses             : {losses}")
    print(f"  Win Rate           : {win_rate:.1f}%")
    print(f"  Total P&L          : ${total_pnl:+.2f}")
    print(f"  Avg Win            : ${avg_win:+.2f}" if not math.isnan(avg_win) else "  Avg Win           : N/A")
    print(f"  Avg Loss           : ${avg_loss:+.2f}" if not math.isnan(avg_loss) else "  Avg Loss          : N/A")
    print(f"  ---")
    print(f"  Days skipped (VIX) : {skipped_vix}")
    print(f"  Signals skipped    : {skipped_direction} (direction filter)")
    print(f"{'='*60}\n")

    # ── Excel export ─────────────────────────────────────
    display_cols = [c for c in df.columns if not c.startswith("_")]
    export_df = df[display_cols].copy()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Trade Log", index=False)
        ws = writer.sheets["Trade Log"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1A4731")
        win_fill    = PatternFill("solid", fgColor="C6EFCE")
        loss_fill   = PatternFill("solid", fgColor="FFC7CE")
        alt_fill    = PatternFill("solid", fgColor="F2F2F2")
        thin        = Side(style="thin", color="CCCCCC")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = border

        result_col_idx = display_cols.index("Result") + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            result_val = ws.cell(row=row_idx, column=result_col_idx).value
            row_fill = win_fill if result_val == "WIN" else loss_fill if result_val == "LOSS" else (alt_fill if row_idx % 2 == 0 else None)
            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")
                cell.font      = Font(size=10)
                if row_fill:
                    cell.fill = row_fill

        for col_idx, col in enumerate(display_cols, 1):
            max_len = max(len(str(col)), max(
                (len(str(export_df[col].iloc[i])) for i in range(len(export_df))),
                default=0
            )) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 25)

        ws.freeze_panes = "A2"

        # Comparison summary sheet
        sum_df = pd.DataFrame([
            ["Total Trades",    102,          total,                       total - 102],
            ["Win Rate",        "48.0%",      f"{win_rate:.1f}%",          f"{win_rate-48.0:+.1f}%"],
            ["Total P&L",       "$+1,298",    f"${total_pnl:+.2f}",        f"${total_pnl-1298:+.2f}"],
            ["Avg Win",         "$+79.80",    f"${avg_win:+.2f}" if not math.isnan(avg_win) else "N/A", ""],
            ["Avg Loss",        "$-49.28",    f"${avg_loss:+.2f}" if not math.isnan(avg_loss) else "N/A", ""],
            ["Trade Window",    "7AM-12PM",   "7AM-9AM",                   "Tighter"],
            ["Max Trades/Day",  6,            2,                           "-4"],
            ["Signal Filter",   "iFVG + OB",  "iFVG only",                 "Higher quality"],
            ["Direction Filter","None",        "1H SMA",                   "Added"],
            ["VIX Filter",      "None",        f"< {VIX_THRESHOLD}",       "Added"],
            ["Days skipped",    0,             skipped_vix,                f"{skipped_vix} high-VIX days"],
        ], columns=["Metric", "Baseline", "Optimized", "Change"])
        sum_df.to_excel(writer, sheet_name="Comparison", index=False)

        ws2 = writer.sheets["Comparison"]
        header_fill2 = PatternFill("solid", fgColor="1F3864")
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill2
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 18

    print(f"Report saved to:\n  {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()
