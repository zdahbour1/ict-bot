"""
ICT Strategy Backtest — 1-MINUTE BARS
======================================
Uses 1-minute bars (same as live bot) for maximum signal precision.
Limited to last 7 days (yfinance free limit for 1m data).
Generates Excel report.
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import math
import logging
import warnings
import pandas as pd
import numpy as np
import yfinance as yf
import pytz
from datetime import datetime, timedelta, date

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.WARNING)

from strategy.ict_long  import run_strategy as run_long
from strategy.ict_short import run_strategy_short as run_short
from strategy.levels    import get_all_levels
from data.aggregator    import aggregate

TICKER         = "QQQ"
CONTRACTS      = 2
PROFIT_TARGET  = 0.25
STOP_LOSS      = 0.15
MAX_TRADES_DAY = 6
TRADE_START_H  = 7
TRADE_END_H    = 12
PT             = pytz.timezone("America/Los_Angeles")
OUTPUT_FILE    = os.path.join(os.path.dirname(__file__), "ICT_Backtest_1m.xlsx")


def bar_to_pt(ts):
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    return ts.tz_convert(PT)


def estimate_option_price(qqq_price, bars_1m, entry_bar):
    lookback = min(60, entry_bar)
    recent = bars_1m.iloc[max(0, entry_bar - lookback):entry_bar]
    if len(recent) < 2:
        iv = 0.22
    else:
        rets = np.log(recent["close"] / recent["close"].shift(1)).dropna()
        iv = rets.std() * math.sqrt(252 * 390)   # 390 1m bars per day
        iv = max(iv, 0.10)
    T = 3.0 / (252 * 6.5)
    price = qqq_price * iv * math.sqrt(T) * 0.4
    return max(round(price, 2), 0.05)


def simulate_exit(bars_1m, entry_bar, entry_opt_px, direction):
    tp_px        = entry_opt_px * (1 + PROFIT_TARGET)
    sl_pct       = -STOP_LOSS
    peak_pnl_pct = 0.0
    entry_close  = bars_1m.iloc[entry_bar]["close"]
    MAX_BARS     = 90   # 90 minutes on 1m bars

    for i in range(entry_bar + 1, min(entry_bar + 500, len(bars_1m))):
        bar      = bars_1m.iloc[i]
        bar_time = bars_1m.index[i]
        bar_pt   = bar_to_pt(bar_time)
        bars_in  = i - entry_bar

        underlying_chg = bar["close"] - entry_close
        opt_chg = underlying_chg * 0.5 if direction == "LONG" else -underlying_chg * 0.5
        cur_opt = max(round(entry_opt_px + opt_chg, 2), 0.01)
        pnl_pct = (cur_opt - entry_opt_px) / entry_opt_px

        if pnl_pct > peak_pnl_pct:
            peak_pnl_pct = pnl_pct

        if peak_pnl_pct >= 0.20:
            sl_pct = peak_pnl_pct - 0.10
        elif peak_pnl_pct >= 0.10:
            sl_pct = 0.00

        hit_tp    = pnl_pct >= PROFIT_TARGET
        hit_sl    = pnl_pct <= sl_pct
        eod       = bar_pt.hour >= 13
        time_exit = bars_in >= MAX_BARS

        if hit_tp or hit_sl or eod or time_exit:
            if hit_tp:
                exit_px = round(tp_px, 2)
                pnl_pct = PROFIT_TARGET
                result  = "WIN"
                reason  = "TP"
            elif hit_sl and sl_pct == 0.0:
                exit_px = entry_opt_px
                pnl_pct = 0.0
                result  = "SCRATCH"
                reason  = "BREAKEVEN"
            elif hit_sl:
                exit_px = max(round(entry_opt_px * (1 + sl_pct), 2), 0.01)
                result  = "WIN" if sl_pct > 0 else "LOSS"
                reason  = "TRAIL SL" if sl_pct > 0 else "SL"
            elif time_exit:
                exit_px = cur_opt
                result  = "WIN" if pnl_pct > 0 else "LOSS" if pnl_pct < 0 else "SCRATCH"
                reason  = "TIME EXIT (90min)"
            else:
                exit_px = cur_opt
                result  = "WIN" if pnl_pct > 0 else "LOSS"
                reason  = "EOD"

            pnl_usd = round((exit_px - entry_opt_px) * 100 * CONTRACTS, 2)
            return {
                "exit_time":      bar_time,
                "exit_option_px": exit_px,
                "pnl_pct":        round(pnl_pct * 100, 1),
                "pnl_usd":        pnl_usd,
                "result":         result,
                "exit_reason":    reason,
                "bars_held":      bars_in,
            }

    last = bars_1m.iloc[min(entry_bar + 499, len(bars_1m) - 1)]
    chg  = last["close"] - entry_close
    opt_chg = chg * 0.5 if direction == "LONG" else -chg * 0.5
    cur  = max(round(entry_opt_px + opt_chg, 2), 0.01)
    pnl  = (cur - entry_opt_px) / entry_opt_px
    return {
        "exit_time":      bars_1m.index[min(entry_bar + 499, len(bars_1m) - 1)],
        "exit_option_px": cur,
        "pnl_pct":        round(pnl * 100, 1),
        "pnl_usd":        round((cur - entry_opt_px) * 100 * CONTRACTS, 2),
        "result":         "WIN" if pnl > 0 else "LOSS",
        "exit_reason":    "END",
        "bars_held":      499,
    }


def occ_symbol(ticker, exp_date, direction, strike):
    opt_type   = "C" if direction == "LONG" else "P"
    strike_int = int(round(strike * 1000))
    return f"{ticker}{exp_date.strftime('%y%m%d')}{opt_type}{strike_int:08d}"


def main():
    print("=" * 60)
    print("  ICT Strategy Backtest — 1-MINUTE BARS")
    print("  Period: Last 7 days (yfinance 1m limit)")
    print("  Trailing stop + 90-min time exit included")
    print("=" * 60)
    print()

    print("Downloading QQQ 1-minute data...")
    raw_1m = yf.download(TICKER, period="7d", interval="1m",
                         auto_adjust=True, progress=False)

    if raw_1m.empty:
        print("ERROR: Could not download data.")
        return

    if isinstance(raw_1m.columns, pd.MultiIndex):
        raw_1m.columns = raw_1m.columns.get_level_values(0).str.lower()
    else:
        raw_1m.columns = raw_1m.columns.str.lower()

    if raw_1m.index.tzinfo is None:
        raw_1m.index = raw_1m.index.tz_localize("UTC")
    raw_1m.index = raw_1m.index.tz_convert(PT)

    bars_1m = raw_1m.copy()
    bars_1h = aggregate(bars_1m, "1h")
    bars_4h = aggregate(bars_1m, "4h")

    trading_days = sorted(set(bars_1m.index.date))
    total_bars   = len(bars_1m)
    print(f"Data loaded: {len(trading_days)} trading days | {total_bars:,} one-minute bars\n")

    trades = []

    for day_date in trading_days:
        day_end  = pd.Timestamp(day_date, tz=PT) + pd.Timedelta(hours=23, minutes=59)
        hist_1m  = bars_1m[bars_1m.index <= day_end].copy()
        hist_1h  = bars_1h[bars_1h.index <= day_end].copy()
        hist_4h  = bars_4h[bars_4h.index <= day_end].copy()

        if len(hist_1m) < 60:
            continue

        day_1m = bars_1m[bars_1m.index.date == day_date]
        if day_1m.empty:
            continue

        try:
            levels = get_all_levels(hist_1m, hist_1h, hist_4h)
        except Exception:
            continue

        if not levels:
            continue

        scan_bars = hist_1m.iloc[-400:]

        try:
            long_signals  = run_long(scan_bars,  hist_1h, hist_4h, levels)
            short_signals = run_short(scan_bars, hist_1h, hist_4h, levels)
        except Exception as e:
            print(f"  {day_date}  strategy error: {e}")
            continue

        all_signals = []
        for s in long_signals:
            s["direction"] = "LONG"
            all_signals.append(s)
        for s in short_signals:
            s["direction"] = "SHORT"
            all_signals.append(s)

        all_signals.sort(key=lambda x: x.get("entry_time", pd.Timestamp.min))

        trades_today  = 0
        last_exit_bar = -1

        for sig in all_signals:
            if trades_today >= MAX_TRADES_DAY:
                break

            entry_time = sig.get("entry_time")
            if entry_time is None:
                continue

            et_pt = bar_to_pt(entry_time) if entry_time.tzinfo is None else entry_time.tz_convert(PT)
            if not (TRADE_START_H <= et_pt.hour < TRADE_END_H):
                continue

            entry_bar = sig.get("entry_bar", 0)
            if entry_bar <= last_exit_bar:
                continue

            entry_price = sig.get("entry_price", 0)
            if entry_price <= 0:
                continue

            direction  = sig["direction"]
            strike     = round(entry_price)
            opt_px     = estimate_option_price(entry_price, scan_bars, entry_bar)
            symbol     = occ_symbol(TICKER, day_date, direction, strike)
            total_cost = round(opt_px * 100 * CONTRACTS, 2)

            exit_info     = simulate_exit(scan_bars, entry_bar, opt_px, direction)
            last_exit_bar = exit_info.get("bars_held", 0) + entry_bar

            exit_time  = exit_info["exit_time"]
            et_pt_exit = bar_to_pt(exit_time) if exit_time.tzinfo is None else exit_time.tz_convert(PT)

            trades.append({
                "Date":             day_date.strftime("%Y-%m-%d"),
                "Direction":        direction,
                "Signal Type":      sig.get("signal_type", ""),
                "Raided Level":     sig.get("raid", {}).get("raided_level", ""),
                "Entry Time (PT)":  et_pt.strftime("%I:%M %p"),
                "QQQ Entry Price":  f"${entry_price:.2f}",
                "Option Symbol":    symbol,
                "Strike":           f"${strike}",
                "Option Entry $":   f"${opt_px:.2f}",
                "Contracts":        CONTRACTS,
                "Total Cost":       f"${total_cost:.2f}",
                "SL (QQQ)":         f"${sig.get('sl', 0):.2f}",
                "TP (QQQ)":         f"${sig.get('tp', 0):.2f}",
                "Exit Time (PT)":   et_pt_exit.strftime("%I:%M %p"),
                "Option Exit $":    f"${exit_info['exit_option_px']:.2f}",
                "Exit Reason":      exit_info.get("exit_reason", ""),
                "P&L %":            f"{exit_info['pnl_pct']:+.1f}%",
                "P&L $":            f"${exit_info['pnl_usd']:+.2f}",
                "Result":           exit_info["result"],
                "_pnl_usd_raw":     exit_info["pnl_usd"],
                "_result_raw":      exit_info["result"],
            })
            trades_today += 1

        if trades_today > 0:
            print(f"  {day_date}  ->  {trades_today} trade(s)")

    if not trades:
        print("\nNo signals found.")
        return

    df       = pd.DataFrame(trades)
    wins     = (df["_result_raw"] == "WIN").sum()
    losses   = (df["_result_raw"] == "LOSS").sum()
    scratches= (df["_result_raw"] == "SCRATCH").sum()
    total    = len(df)
    win_rate = wins / (wins + losses) * 100 if (wins + losses) > 0 else 0
    total_pnl = df["_pnl_usd_raw"].sum()
    avg_win   = df.loc[df["_result_raw"] == "WIN",  "_pnl_usd_raw"].mean()
    avg_loss  = df.loc[df["_result_raw"] == "LOSS", "_pnl_usd_raw"].mean()

    print(f"\n{'='*60}")
    print(f"  1-MINUTE BACKTEST RESULTS (7 days)")
    print(f"{'='*60}")
    print(f"  Total Trades : {total}")
    print(f"  Wins         : {wins}")
    print(f"  Losses       : {losses}")
    print(f"  Scratches    : {scratches}  (breakeven exits)")
    print(f"  Win Rate     : {win_rate:.1f}%  (excl. scratches)")
    print(f"  Total P&L    : ${total_pnl:+.2f}")
    print(f"  Avg Win      : ${avg_win:+.2f}" if not math.isnan(avg_win) else "  Avg Win      : N/A")
    print(f"  Avg Loss     : ${avg_loss:+.2f}" if not math.isnan(avg_loss) else "  Avg Loss     : N/A")
    print(f"{'='*60}")
    print(f"\n  NOTE: 7-day sample is too small for strong conclusions.")
    print(f"  Run again in 2-3 weeks to build a larger 1m dataset.\n")

    # Excel export
    display_cols = [c for c in df.columns if not c.startswith("_")]
    export_df    = df[display_cols].copy()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Trade Log (1m)", index=False)
        ws = writer.sheets["Trade Log (1m)"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F3864")
        win_fill    = PatternFill("solid", fgColor="C6EFCE")
        loss_fill   = PatternFill("solid", fgColor="FFC7CE")
        scratch_fill= PatternFill("solid", fgColor="FFEB9C")
        thin        = Side(style="thin", color="CCCCCC")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = border

        result_col_idx = display_cols.index("Result") + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            result_val = ws.cell(row=row_idx, column=result_col_idx).value
            if result_val == "WIN":
                row_fill = win_fill
            elif result_val == "LOSS":
                row_fill = loss_fill
            elif result_val == "SCRATCH":
                row_fill = scratch_fill
            else:
                row_fill = None
            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")
                cell.font      = Font(size=10)
                if row_fill:
                    cell.fill = row_fill

        for col_idx, col in enumerate(display_cols, 1):
            max_len = max(len(str(col)), max(
                (len(str(export_df[col].iloc[i])) for i in range(len(export_df))),
                default=0
            )) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 25)

        ws.freeze_panes = "A2"

        # Summary sheet
        avg_win_str  = f"${avg_win:+.2f}"  if not math.isnan(avg_win)  else "N/A"
        avg_loss_str = f"${avg_loss:+.2f}" if not math.isnan(avg_loss) else "N/A"
        sum_df = pd.DataFrame([
            ["Period",           "Last 7 trading days (1m bars)"],
            ["Ticker",           TICKER],
            ["Total Trades",     total],
            ["Wins",             wins],
            ["Losses",           losses],
            ["Scratches",        scratches],
            ["Win Rate",         f"{win_rate:.1f}% (excl. scratches)"],
            ["Total P&L",        f"${total_pnl:+.2f}"],
            ["Avg Win",          avg_win_str],
            ["Avg Loss",         avg_loss_str],
            ["Profit Target",    "25%"],
            ["Stop Loss",        "15% (trailing)"],
            ["Time Exit",        "90 minutes"],
            ["Contracts",        CONTRACTS],
            ["Trade Window",     "7:00 AM - 12:00 PM PT"],
            ["Bar Timeframe",    "1-minute (same as live bot)"],
            ["", ""],
            ["NOTE", "7-day sample is too small for strong conclusions. "
                     "Re-run in 2-3 weeks for a more meaningful dataset."],
        ], columns=["Metric", "Value"])
        sum_df.to_excel(writer, sheet_name="Summary", index=False)

        ws2 = writer.sheets["Summary"]
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(size=10)
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 55

    print(f"Report saved to:\n  {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()
