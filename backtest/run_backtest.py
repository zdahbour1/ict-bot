"""
ICT Strategy Backtest — QQQ 0DTE Options
=========================================
Runs both LONG and SHORT ICT strategies on historical 5-minute bars.
Data: yfinance (60 days of 5m bars — free limit).
Options: Estimated using simplified Black-Scholes (historical prices unavailable for free).
Generates a formatted Excel report.
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import math
import logging
import warnings
import pandas as pd
import numpy as np
import yfinance as yf
import pytz
from datetime import datetime, timedelta, date

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.WARNING)   # suppress noisy SDK logs

from strategy.ict_long  import run_strategy as run_long
from strategy.ict_short import run_strategy_short as run_short
from strategy.levels    import get_all_levels

# ── Config ─────────────────────────────────────────────────
# Usage: python run_backtest.py [TICKER]   (default: all tickers)
#        python run_backtest.py QQQ
#        python run_backtest.py ALL
import importlib
_config = importlib.import_module("config")
ALL_TICKERS     = getattr(_config, "TICKERS", ["QQQ"])
CONTRACTS       = 2
PROFIT_TARGET   = 1.00      # 100% TP
STOP_LOSS       = 0.60      # 60% SL
TRAIL_BE        = 0.10      # +10% → move stop to breakeven
TRAIL_START     = 0.20      # +20% → start trailing
TRAIL_DIST      = 0.10      # trail 10% below peak
TIME_EXIT_MIN   = 90        # 90 min time exit
MAX_TRADES_DAY  = 999       # no daily limit
TRADE_START_H   = 6         # 6:30 AM PT
TRADE_START_MIN = 30
TRADE_END_H     = 12        # 12 PM PT
PT              = pytz.timezone("America/Los_Angeles")
BACKTEST_DIR    = os.path.dirname(__file__)


# ── Helpers ────────────────────────────────────────────────

def bar_to_pt(ts, tz=PT):
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    return ts.tz_convert(tz)


def estimate_option_price(qqq_price: float, bars_5m: pd.DataFrame,
                           entry_bar: int) -> float:
    """
    Estimate ATM 0DTE option price via simplified Black-Scholes.
    Uses realized vol from recent 20 bars as proxy for IV.
    """
    lookback = min(20, entry_bar)
    recent = bars_5m.iloc[max(0, entry_bar - lookback):entry_bar]
    if len(recent) < 2:
        iv = 0.22
    else:
        rets = np.log(recent["close"] / recent["close"].shift(1)).dropna()
        iv = rets.std() * math.sqrt(252 * 78)     # 78 five-min bars per day
        iv = max(iv, 0.10)

    # Assume 3 hours remaining on average (mid-morning entry)
    T = 3.0 / (252 * 6.5)
    price = qqq_price * iv * math.sqrt(T) * 0.4   # ATM approximation
    return max(round(price, 2), 0.05)


def simulate_exit(bars_5m: pd.DataFrame, entry_bar: int,
                  entry_opt_px: float, direction: str):
    """
    Walk forward bar-by-bar after entry.
    Option P&L estimated using delta≈0.5 for ATM.
    Trailing stop: +10% → breakeven, +20% → trail at peak - 10%
    TP: +100%, SL: -60%, Time exit: 90 min, EOD: 1 PM PT
    """
    entry_close  = bars_5m.iloc[entry_bar]["close"]
    peak_pnl_pct = 0.0
    dynamic_sl   = -STOP_LOSS
    MAX_BARS     = TIME_EXIT_MIN // 5  # 90min / 5min bars = 18 bars

    for i in range(entry_bar + 1, min(entry_bar + 300, len(bars_5m))):
        bar      = bars_5m.iloc[i]
        bar_time = bars_5m.index[i]
        bar_pt   = bar_to_pt(bar_time)
        bars_in  = i - entry_bar

        # Underlying move → option P&L (delta ~0.5 ATM)
        underlying_chg = bar["close"] - entry_close
        opt_chg = underlying_chg * 0.5 if direction == "LONG" else -underlying_chg * 0.5
        cur_opt = max(round(entry_opt_px + opt_chg, 2), 0.01)
        pnl_pct = (cur_opt - entry_opt_px) / entry_opt_px

        # No trailing stop — fixed TP and SL only
        hit_tp = pnl_pct >= PROFIT_TARGET
        hit_sl = pnl_pct <= -STOP_LOSS
        time_exit = bars_in >= MAX_BARS
        eod       = bar_pt.hour >= 13

        if hit_tp or hit_sl or time_exit or eod:
            if hit_tp:
                result = "WIN"
                reason = "TAKE PROFIT"
                cur_opt = round(entry_opt_px * (1 + PROFIT_TARGET), 2)
            elif hit_sl:
                result = "LOSS"
                reason = "STOP LOSS"
            elif time_exit:
                result = "WIN" if pnl_pct > 0 else "LOSS"
                reason = "TIME EXIT"
            else:
                result = "WIN" if pnl_pct > 0 else "LOSS"
                reason = "EOD EXIT"

            pnl_usd = round((cur_opt - entry_opt_px) * 100 * CONTRACTS, 2)
            return {
                "exit_time":       bar_time,
                "exit_option_px":  cur_opt,
                "pnl_pct":         round(pnl_pct * 100, 1),
                "pnl_usd":         pnl_usd,
                "result":          result,
                "reason":          reason,
                "bars_held":       bars_in,
            }

    # Fallback: last bar
    last = bars_5m.iloc[min(entry_bar + 299, len(bars_5m) - 1)]
    chg  = last["close"] - entry_close
    opt_chg = chg * 0.5 if direction == "LONG" else -chg * 0.5
    cur  = max(round(entry_opt_px + opt_chg, 2), 0.01)
    pnl  = (cur - entry_opt_px) / entry_opt_px
    return {
        "exit_time":      bars_5m.index[min(entry_bar + 299, len(bars_5m) - 1)],
        "exit_option_px": cur,
        "pnl_pct":        round(pnl * 100, 1),
        "pnl_usd":        round((cur - entry_opt_px) * 100 * CONTRACTS, 2),
        "result":         "WIN" if pnl > 0 else "LOSS",
        "bars_held":      299,
    }


def occ_symbol(ticker: str, exp_date: date, direction: str, strike: float) -> str:
    """Build OCC-style option symbol. e.g. QQQ260327C00567000"""
    opt_type = "C" if direction == "LONG" else "P"
    strike_int = int(round(strike * 1000))
    return f"{ticker}{exp_date.strftime('%y%m%d')}{opt_type}{strike_int:08d}"


# ── Main backtest ──────────────────────────────────────────

def run_backtest_for_ticker(TICKER):
    """Run backtest for a single ticker. Returns (ticker, summary_dict, trades_df) or None."""
    OUTPUT_FILE = os.path.join(BACKTEST_DIR, f"ICT_Backtest_Report_{TICKER}.xlsx")

    print(f"\n{'='*60}")
    print(f"  ICT Strategy Backtest — {TICKER} 0DTE Options")
    print(f"  Period: Last 60 trading days (max free data)")
    print(f"  Strategy: LONG + SHORT  |  100% TP / 60% SL / Trail +10% BE / +20% trail")
    print(f"{'='*60}")
    print()

    # ── Fetch data ───────────────────────────────────────
    print(f"Downloading {TICKER} data from yfinance...")
    raw_5m = yf.download(TICKER, period="60d", interval="5m",
                         auto_adjust=True, progress=False)
    raw_1h = yf.download(TICKER, period="60d", interval="1h",
                         auto_adjust=True, progress=False)
    raw_4h = yf.download(TICKER, period="60d", interval="1h",
                         auto_adjust=True, progress=False)   # resample to 4h

    if raw_5m.empty:
        print(f"ERROR: Could not download data for {TICKER}. Skipping.")
        return None

    # Flatten MultiIndex columns if present
    for df_ref in [raw_5m, raw_1h, raw_4h]:
        if isinstance(df_ref.columns, pd.MultiIndex):
            df_ref.columns = df_ref.columns.get_level_values(0).str.lower()
        else:
            df_ref.columns = df_ref.columns.str.lower()

    # Resample 1h → 4h
    raw_4h = raw_4h.resample("4h").agg({
        "open": "first", "high": "max",
        "low": "min",  "close": "last",
        "volume": "sum"
    }).dropna()

    # Localize / convert to PT
    def to_pt(df):
        if df.index.tzinfo is None:
            df.index = df.index.tz_localize("UTC")
        df.index = df.index.tz_convert(PT)
        return df

    bars_5m = to_pt(raw_5m.copy())
    bars_1h = to_pt(raw_1h.copy())
    bars_4h = to_pt(raw_4h.copy())

    # Rename columns to lowercase
    for df in [bars_5m, bars_1h, bars_4h]:
        df.columns = [c.lower() for c in df.columns]

    trading_days = sorted(set(bars_5m.index.date))
    print(f"Data loaded: {len(trading_days)} trading days  |  "
          f"{len(bars_5m):,} five-minute bars\n")

    # ── Day-by-day backtest ──────────────────────────────
    trades = []

    for day_date in trading_days:
        # Slice data up to end of this day (no lookahead)
        day_end = pd.Timestamp(day_date) + pd.Timedelta(hours=23, minutes=59)
        day_end = day_end.tz_localize(PT)

        hist_5m = bars_5m[bars_5m.index <= day_end].copy()
        hist_1h = bars_1h[bars_1h.index <= day_end].copy()
        hist_4h = bars_4h[bars_4h.index <= day_end].copy()

        if len(hist_5m) < 50:
            continue

        # Only bars within this trading day (7 AM - 1 PM PT)
        day_5m = bars_5m[bars_5m.index.date == day_date]
        if day_5m.empty:
            continue

        # Compute levels using data up to start of this day
        try:
            levels = get_all_levels(hist_5m, hist_1h, hist_4h)
        except Exception:
            continue

        if not levels:
            continue

        # Run strategy on this day's bars
        try:
            long_signals  = run_long(day_5m,  hist_1h, hist_4h, levels)
            short_signals = run_short(day_5m, hist_1h, hist_4h, levels)
        except Exception:
            continue

        all_signals = []
        for s in long_signals:
            s["direction"] = "LONG"
            all_signals.append(s)
        for s in short_signals:
            s["direction"] = "SHORT"
            all_signals.append(s)

        # Sort by entry time
        all_signals.sort(key=lambda x: x.get("entry_time", pd.Timestamp.min))

        trades_today = 0
        last_exit_bar = -1

        for sig in all_signals:
            if trades_today >= MAX_TRADES_DAY:
                break

            entry_time = sig.get("entry_time")
            if entry_time is None:
                continue

            # Only within trade window (6:30 AM - 12:00 PM PT)
            et_pt = bar_to_pt(entry_time)
            in_window = (
                (et_pt.hour > TRADE_START_H or
                 (et_pt.hour == TRADE_START_H and et_pt.minute >= TRADE_START_MIN))
                and et_pt.hour < TRADE_END_H
            )
            if not in_window:
                continue

            entry_bar = sig.get("entry_bar", 0)

            # Don't enter while previous trade is still open
            if entry_bar <= last_exit_bar:
                continue

            entry_price = sig.get("entry_price", 0)
            if entry_price <= 0:
                continue

            direction = sig["direction"]
            strike    = round(entry_price)   # ATM strike
            opt_px    = estimate_option_price(entry_price, day_5m, entry_bar)
            symbol    = occ_symbol(TICKER, day_date, direction,
                                   strike)
            total_cost = round(opt_px * 100 * CONTRACTS, 2)

            # Simulate exit
            exit_info = simulate_exit(day_5m, entry_bar, opt_px, direction)
            last_exit_bar = exit_info.get("bars_held", 0) + entry_bar

            exit_time = exit_info["exit_time"]
            et_pt_exit = bar_to_pt(exit_time)

            trades.append({
                "Date":             day_date.strftime("%Y-%m-%d"),
                "Direction":        direction,
                "Signal Type":      sig.get("signal_type", ""),
                "Raided Level":     sig.get("raid", {}).get("raided_level", ""),
                "Entry Time (PT)":  et_pt.strftime("%I:%M %p"),
                "Entry Price":      f"${entry_price:.2f}",
                "Option Symbol":    symbol,
                "Strike":           f"${strike}",
                "Option Entry $":   f"${opt_px:.2f}",
                "Contracts":        CONTRACTS,
                "Total Cost":       f"${total_cost:.2f}",
                "SL (QQQ)":         f"${sig.get('sl', 0):.2f}",
                "TP (QQQ)":         f"${sig.get('tp', 0):.2f}",
                "Exit Time (PT)":   et_pt_exit.strftime("%I:%M %p"),
                "Option Exit $":    f"${exit_info['exit_option_px']:.2f}",
                "P&L %":            f"{exit_info['pnl_pct']:+.1f}%",
                "P&L $":            f"${exit_info['pnl_usd']:+.2f}",
                "Result":           exit_info["result"],
                "Exit Reason":      exit_info.get("reason", ""),
                "_pnl_usd_raw":     exit_info["pnl_usd"],
                "_result_raw":      exit_info["result"],
            })
            trades_today += 1

        if trades_today > 0:
            print(f"  {day_date}  ->  {trades_today} trade(s)")

    if not trades:
        print("\nNo signals found in this period.")
        return

    # ── Build summary stats ──────────────────────────────
    df = pd.DataFrame(trades)
    wins   = (df["_result_raw"] == "WIN").sum()
    losses = (df["_result_raw"] == "LOSS").sum()
    total  = len(df)
    win_rate = wins / total * 100 if total > 0 else 0
    total_pnl = df["_pnl_usd_raw"].sum()
    avg_win   = df.loc[df["_result_raw"] == "WIN",  "_pnl_usd_raw"].mean()
    avg_loss  = df.loc[df["_result_raw"] == "LOSS", "_pnl_usd_raw"].mean()

    print(f"\n{'='*60}")
    print(f"  RESULTS SUMMARY")
    print(f"{'='*60}")
    print(f"  Total Trades : {total}")
    print(f"  Wins         : {wins}")
    print(f"  Losses       : {losses}")
    print(f"  Win Rate     : {win_rate:.1f}%")
    print(f"  Total P&L    : ${total_pnl:+.2f}")
    print(f"  Avg Win      : ${avg_win:+.2f}" if not math.isnan(avg_win) else "  Avg Win      : N/A")
    print(f"  Avg Loss     : ${avg_loss:+.2f}" if not math.isnan(avg_loss) else "  Avg Loss     : N/A")
    print(f"{'='*60}\n")

    # ── Export to Excel ──────────────────────────────────
    display_cols = [c for c in df.columns if not c.startswith("_")]
    export_df = df[display_cols].copy()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Trade log sheet
        export_df.to_excel(writer, sheet_name="Trade Log", index=False)
        ws = writer.sheets["Trade Log"]

        # Header formatting
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F3864")
        win_fill    = PatternFill("solid", fgColor="C6EFCE")
        loss_fill   = PatternFill("solid", fgColor="FFC7CE")
        alt_fill    = PatternFill("solid", fgColor="F2F2F2")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = border

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            result_col = display_cols.index("Result")
            result_val = ws.cell(row=row_idx, column=result_col + 1).value
            row_fill = win_fill if result_val == "WIN" else loss_fill if result_val == "LOSS" else (alt_fill if row_idx % 2 == 0 else None)

            for cell in row:
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")
                cell.font      = Font(size=10)
                if row_fill:
                    cell.fill = row_fill

        # Auto column widths
        for col_idx, col in enumerate(display_cols, 1):
            max_len = max(len(str(col)), max(
                (len(str(export_df[col].iloc[i])) for i in range(len(export_df))),
                default=0
            )) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 25)

        ws.freeze_panes = "A2"

        # Summary sheet
        summary_data = {
            "Metric": [
                "Period", "Ticker", "Total Trades", "Wins", "Losses",
                "Win Rate", "Total P&L", "Avg Win", "Avg Loss",
                "Profit Target", "Stop Loss", "Contracts per Trade",
                "Trade Window", "Max Trades/Day",
                "", "DISCLAIMER",
            ],
            "Value": [
                f"Last 60 trading days", TICKER,
                total, wins, losses,
                f"{win_rate:.1f}%",
                f"${total_pnl:+.2f}",
                f"${avg_win:+.2f}" if not math.isnan(avg_win) else "N/A",
                f"${avg_loss:+.2f}" if not math.isnan(avg_loss) else "N/A",
                "25%", "15%", CONTRACTS,
                "7:00 AM – 12:00 PM PT",
                MAX_TRADES_DAY,
                "",
                "Option prices are ESTIMATED (historical 0DTE prices unavailable for free). "
                "Real results will vary. This is for educational purposes only.",
            ]
        }
        sum_df = pd.DataFrame(summary_data)
        sum_df.to_excel(writer, sheet_name="Summary", index=False)

        ws2 = writer.sheets["Summary"]
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(size=10)
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 70

    print(f"[{TICKER}] Excel report saved to:\n  {OUTPUT_FILE}\n")

    return {
        "ticker": TICKER,
        "total_trades": total,
        "wins": wins,
        "losses": losses,
        "win_rate": win_rate,
        "total_pnl": total_pnl,
        "avg_win": avg_win if not math.isnan(avg_win) else 0,
        "avg_loss": avg_loss if not math.isnan(avg_loss) else 0,
    }


def main():
    """Run backtest for one or all tickers. Usage: python run_backtest.py [TICKER|ALL]"""
    arg = sys.argv[1].upper() if len(sys.argv) > 1 else "ALL"

    if arg == "ALL":
        tickers = ALL_TICKERS
    else:
        tickers = [arg]

    results = []
    for ticker in tickers:
        result = run_backtest_for_ticker(ticker)
        if result:
            results.append(result)

    if len(results) > 1:
        # Print comparison summary
        print(f"\n{'='*70}")
        print(f"  MULTI-TICKER COMPARISON SUMMARY")
        print(f"{'='*70}")
        print(f"  {'Ticker':<8} {'Trades':>8} {'Wins':>6} {'Losses':>8} {'Win %':>8} {'Total P&L':>12} {'Avg Win':>10} {'Avg Loss':>10}")
        print(f"  {'-'*62}")

        for r in sorted(results, key=lambda x: x["total_pnl"], reverse=True):
            print(f"  {r['ticker']:<8} {r['total_trades']:>8} {r['wins']:>6} {r['losses']:>8} "
                  f"{r['win_rate']:>7.1f}% ${r['total_pnl']:>+10.2f} ${r['avg_win']:>+8.2f} ${r['avg_loss']:>+8.2f}")

        best = max(results, key=lambda x: x["total_pnl"])
        print(f"\n  Most profitable ticker: {best['ticker']} (${best['total_pnl']:+.2f})")
        print(f"{'='*70}\n")

        # Save comparison CSV
        comp_path = os.path.join(BACKTEST_DIR, "backtest_comparison.csv")
        comp_df = pd.DataFrame(results)
        comp_df = comp_df.sort_values("total_pnl", ascending=False)
        comp_df.to_csv(comp_path, index=False)
        print(f"  Comparison saved to: {comp_path}")


if __name__ == "__main__":
    main()
